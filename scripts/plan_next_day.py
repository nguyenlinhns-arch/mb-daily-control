#!/usr/bin/env python3
"""Build the next-day MB review and publishable website payload.

This module is intentionally deterministic and idempotent:
- Reads the latest locked 27-code history from the exported Google workbook.
- Recomputes A1 Core100 / Volume50, X3 Growth32-34, and guarded X2 Rescue.
- Enforces one real-money order maximum with controller order:
  A1 Core -> A1 Volume -> X3 Growth -> X2 Rescue -> A0.
- Writes data/current.json plus a dated immutable plan snapshot.
- Every displayed A1/X2/X3 candidate carries an earliest eligible date and
  condition. These dates are lower bounds and are recalculated after each lock.
- Never includes X2 performance/backtest metrics in the website payload.
- Never records P/L unless a separately confirmed real order is settled.
"""
from __future__ import annotations

import argparse
import copy
import hashlib
import json
import math
import os
import re
import tempfile
import urllib.request
from collections import Counter
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DATA_FILE = ROOT / "data" / "current.json"
PLAN_LEDGER_FILE = ROOT / "data" / "review-ledger.json"
AUTOMATION_STATE_FILE = ROOT / "data" / "automation-state.json"
PLANS_DIR = ROOT / "data" / "plans"
POLICY_FILE = ROOT / "data" / "automation-policy.json"

SHEET_ID = os.getenv(
    "GOOGLE_SHEET_ID", "1iVAfqmS-TvP02U8FtKSM2nr_7Dsd7qi2qEGnWV6IK7w"
)
EXPORT_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"
SOURCE_XLSX_PATH = os.getenv("SOURCE_XLSX_PATH")
VN = timezone(timedelta(hours=7))

# Latest user-approved operating standard for the website automation.
A1_CORE_POINTS = 100
A1_VOLUME_POINTS = 50
LO_COST_PER_POINT = 23_000
LO_PAYOUT_PER_HIT_POINT = 80_000
X3_POINTS = 50
X2_POINTS = 15

CONTROLLER_ORDER = ["A1_CORE100", "A1_VOLUME50", "X3_GROWTH", "X2_RESCUE", "A0"]
HISTORY_SHEETS = (
    "MB_History_27",
    "MB_History_27_IMPORT",
    "Lo_Toan_Bang_2024_2026",
    "Raw_Results_2024_2026",
    "Raw_Results_IMPORT",
    "Raw_2Digits_IMPORT",
)
MILESTONE_GROUPS = ("A1", "X2", "X3")


def now_vn() -> datetime:
    return datetime.now(VN)


def iso_day(value: date) -> str:
    return value.isoformat()


def parse_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            return (datetime(1899, 12, 30) + timedelta(days=float(value))).date()
        except Exception:
            return None
    text = str(value or "").strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def code2(value: Any) -> str | None:
    if value is None or value == "":
        return None
    try:
        text = str(int(float(value)))
    except (TypeError, ValueError):
        text = str(value).strip()
    digits = "".join(ch for ch in text if ch.isdigit())
    return digits[-2:].zfill(2) if digits else None


def load_json(path: Path, default: Any) -> Any:
    if not path.exists():
        return copy.deepcopy(default)
    return json.loads(path.read_text(encoding="utf-8"))


def canonical_json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def json_hash(value: Any) -> str:
    return hashlib.sha256(canonical_json(value).encode("utf-8")).hexdigest()


def write_json_if_changed(path: Path, value: Any) -> bool:
    text = json.dumps(value, ensure_ascii=False, separators=(",", ":")) + "\n"
    if path.exists() and path.read_text(encoding="utf-8") == text:
        return False
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")
    return True


def signed_vnd(value: int) -> str:
    return f"{value:+,}đ".replace(",", ".")


def vnd(value: int) -> str:
    return f"{value:,}đ".replace(",", ".")


def find_group(doc: dict[str, Any], group_id: str) -> dict[str, Any] | None:
    for group in doc.get("groups") or []:
        if str(group.get("id", "")).upper() == group_id.upper():
            return group
    return None


def replace_group(doc: dict[str, Any], new_group: dict[str, Any]) -> None:
    group_id = str(new_group.get("id", "")).upper()
    groups = list(doc.get("groups") or [])
    replaced = False
    for idx, group in enumerate(groups):
        if str(group.get("id", "")).upper() == group_id:
            groups[idx] = new_group
            replaced = True
            break
    if not replaced:
        groups.append(new_group)
    order = {"A1": 0, "X2": 1, "X3": 2, "XIEN": 3, "DE": 4}
    groups.sort(key=lambda g: (order.get(str(g.get("id", "")).upper(), 99), str(g.get("id", ""))))
    doc["groups"] = groups


def obtain_xlsx() -> tuple[str, tempfile.NamedTemporaryFile[Any] | None]:
    if SOURCE_XLSX_PATH:
        return SOURCE_XLSX_PATH, None
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx")
    request = urllib.request.Request(EXPORT_URL, headers={"User-Agent": "MB-Daily-Control/3.0"})
    with urllib.request.urlopen(request, timeout=90) as response:
        tmp.write(response.read())
        tmp.flush()
    return tmp.name, tmp


def parse_sheet_history(ws: Any) -> list[tuple[date, list[str]]]:
    by_date: dict[date, list[str]] = {}
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        draw_date = parse_date(row[0])
        codes = [code2(value) for value in row[1:28]]
        if draw_date is None or len(codes) != 27 or not all(codes):
            continue
        clean = [str(value) for value in codes]
        old = by_date.get(draw_date)
        if old is not None and old != clean:
            raise RuntimeError(f"Lệch hai bộ 27 mã trong sheet {ws.title} ngày {draw_date}")
        by_date[draw_date] = clean
    return sorted(by_date.items(), key=lambda item: item[0])


def load_history_and_brake(
    xlsx_path: str, doc: dict[str, Any]
) -> tuple[list[tuple[date, list[str]]], str, dict[str, Any]]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        candidates: list[tuple[int, date, str, list[tuple[date, list[str]]]]] = []
        for name in HISTORY_SHEETS:
            if name not in wb.sheetnames:
                continue
            rows = parse_sheet_history(wb[name])
            if rows:
                candidates.append((len(rows), rows[-1][0], name, rows))
        if not candidates:
            raise RuntimeError("Không tìm thấy lịch sử hợp lệ đủ 27 mã")

        data = doc.get("data") or {}
        locked_date = parse_date(data.get("locked_through"))
        locked_codes = [str(code2(x)).zfill(2) for x in data.get("latest_27_codes", []) if code2(x)]
        if locked_date is None or len(locked_codes) != 27:
            raise RuntimeError("data/current.json chưa có kỳ khóa đủ 27 mã")

        compatible: list[tuple[int, date, str, list[tuple[date, list[str]]]]] = []
        mismatch_same_day: list[str] = []
        for item in candidates:
            _, _, name, rows = item
            row_map = dict(rows)
            if locked_date in row_map and row_map[locked_date] != locked_codes:
                mismatch_same_day.append(name)
                continue
            compatible.append(item)
        if not compatible:
            raise RuntimeError(
                "Lịch sử lệch kỳ khóa hiện tại trong: " + ", ".join(mismatch_same_day)
            )

        # Prefer the longest compatible history, then latest date, then canonical name order.
        preference = {name: len(HISTORY_SHEETS) - idx for idx, name in enumerate(HISTORY_SHEETS)}
        compatible.sort(
            key=lambda item: (item[0], item[1], preference.get(item[2], 0)), reverse=True
        )
        _, _, source_sheet, history = compatible[0]
        history_map = dict(history)
        if locked_date not in history_map:
            history.append((locked_date, locked_codes))
            history.sort(key=lambda item: item[0])
        elif history_map[locked_date] != locked_codes:
            raise RuntimeError(f"{source_sheet} lệch bộ 27 mã đã khóa ngày {locked_date}")

        # Drop any accidental future rows beyond the current locked date.
        history = [item for item in history if item[0] <= locked_date]
        if not history or history[-1][0] != locked_date:
            raise RuntimeError("Không thể ghép lịch sử đến đúng ngày khóa")
        if len(history) < 100:
            raise RuntimeError("Lịch sử dưới 100 kỳ; không đủ để rà soát")

        brake = read_x2_brake(wb)
        return history, source_sheet, brake
    finally:
        wb.close()


def read_x2_brake(wb: Any) -> dict[str, Any]:
    result = {
        "status": "ACTIVE_GUARDED",
        "brake_remaining": 0,
        "loss_streak": 0,
        "drawdown_vnd": 0,
        "source": "X2_Live_Log_v1",
    }
    if "X2_Live_Log_v1" not in wb.sheetnames:
        result["source"] = "NO_X2_LIVE_LOG"
        return result
    ws = wb["X2_Live_Log_v1"]
    header: dict[str, int] | None = None
    last_values: tuple[Any, ...] | None = None
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        if header is None and str(row[0] or "").strip() == "Date":
            header = {str(value or "").strip(): idx for idx, value in enumerate(row)}
            continue
        if header is not None and parse_date(row[header.get("Date", 0)]) is not None:
            last_values = row
    if header is None or last_values is None:
        return result

    def get(name: str, default: Any = None) -> Any:
        idx = header.get(name)
        return last_values[idx] if idx is not None and idx < len(last_values) else default

    def integer(value: Any, default: int = 0) -> int:
        try:
            return int(float(value))
        except (TypeError, ValueError):
            return default

    result.update(
        {
            "status": str(get("Pilot_Status", "ACTIVE_GUARDED") or "ACTIVE_GUARDED"),
            "brake_remaining": integer(get("Brake_Remaining")),
            "loss_streak": integer(get("Loss_Streak")),
            "drawdown_vnd": integer(get("Drawdown_VND")),
        }
    )
    return result


def history_masks(history: list[tuple[date, list[str]]], code: str, window: int) -> int:
    mask = 0
    for _, codes in history[-window:]:
        mask = (mask << 1) | int(code in set(codes))
    return mask


def rolling_counts(history: list[tuple[date, list[str]]], code: str, window: int) -> int:
    return sum(1 for _, codes in history[-window:] if code in set(codes))


def build_features(history: list[tuple[date, list[str]]]) -> dict[str, dict[str, Any]]:
    n = len(history)
    draw_counters = [Counter(codes) for _, codes in history]
    features: dict[str, dict[str, Any]] = {}
    for number in range(100):
        code = f"{number:02d}"
        positions = [idx for idx, counts in enumerate(draw_counters) if counts.get(code, 0) > 0]
        if positions:
            gan = n - 1 - positions[-1]
            completed = [positions[idx + 1] - positions[idx] - 1 for idx in range(len(positions) - 1)]
            gmax = max(completed) if completed else 0
        else:
            gan = n
            gmax = 0
        score = gan / gmax if gmax > 0 else 0.0
        recent5 = draw_counters[-5:]
        feature = {
            "code": code,
            "gan": gan,
            "gmax": gmax,
            "score": score,
            "presence5": sum(1 for counts in recent5 if counts.get(code, 0) > 0),
            "maxfreq5": max((counts.get(code, 0) for counts in recent5), default=0),
            "hot21": sum(counts.get(code, 0) for counts in draw_counters[-21:]),
            "h5": rolling_counts(history, code, 5),
            "h21": rolling_counts(history, code, 21),
            "h60": rolling_counts(history, code, 60),
            "h90": rolling_counts(history, code, 90),
            "in_latest": draw_counters[-1].get(code, 0) > 0,
            "last5_counts": tuple(counts.get(code, 0) for counts in recent5),
            "mask5": history_masks(history, code, 5),
            "mask21": history_masks(history, code, 21),
        }
        feature["x3_score"] = feature["hot21"] + 2 * math.sqrt(min(gan, 30))
        features[code] = feature
    return features


def a1_gate(state: tuple[int, int, tuple[int, ...]], tier: str, noise_blocked: bool) -> bool:
    gan, gmax, last5 = state
    if noise_blocked or gmax <= 0 or any(value > 0 for value in last5):
        return False
    score = gan / gmax
    if tier == "CORE":
        return gan >= 21 and 0.90 <= score <= 1.60
    return gan >= 12 and 0.70 <= score <= 1.80


def earliest_a1_date(
    feature: dict[str, Any], tier: str, target_date: date, noise_blocked: bool, max_days: int = 120
) -> tuple[date, str, str]:
    initial = (int(feature["gan"]), int(feature["gmax"]), tuple(feature["last5_counts"]))
    states: set[tuple[int, int, tuple[int, ...]]] = {initial}
    for offset in range(max_days + 1):
        blocked = noise_blocked if offset == 0 else False
        if any(a1_gate(state, tier, blocked) for state in states):
            label = "ELIGIBLE_NOW" if offset == 0 else "CONDITIONAL_LOWER_BOUND"
            condition = (
                f"Mốc sớm nhất theo mô phỏng hit/miss; {tier} phải còn đủ Gan/Score, "
                "không xuất hiện 5 kỳ và kỳ trước không kích hoạt phanh nhiễu. Tính lại sau mỗi kỳ khóa."
            )
            return target_date + timedelta(days=offset), condition, label
        next_states: set[tuple[int, int, tuple[int, ...]]] = set()
        for gan, gmax, last5 in states:
            # Miss in the next draw.
            next_states.add((gan + 1, gmax, (*last5[1:], 0)))
            # One hit in the next draw; this completes the current gap.
            next_states.add((0, max(gmax, gan), (*last5[1:], 1)))
        # Keep the state space bounded without losing earliest-gate paths.
        if len(next_states) > 20_000:
            next_states = set(
                sorted(next_states, key=lambda x: (x[0], x[1], x[2]))[:20_000]
            )
        states = next_states
    return (
        target_date + timedelta(days=1),
        f"Chưa có đường vào {tier} cố định trong {max_days} kỳ; mốc hiển thị là lần rà lại sớm nhất sau kỳ khóa kế tiếp.",
        "DYNAMIC_RECALCULATION",
    )


def a1_candidates(
    features: dict[str, dict[str, Any]], repeat2_count: int, max_frequency: int, target_date: date
) -> dict[str, Any]:
    noise_blocked = repeat2_count >= 3 or max_frequency >= 3
    core = [
        f
        for f in features.values()
        if f["gan"] >= 21
        and 0.90 <= f["score"] <= 1.60
        and f["presence5"] == 0
        and f["maxfreq5"] < 2
    ]
    volume = [
        f
        for f in features.values()
        if f["gan"] >= 12
        and 0.70 <= f["score"] <= 1.80
        and f["presence5"] == 0
        and f["maxfreq5"] < 2
    ]
    core.sort(key=lambda f: (-f["gmax"], -f["score"], -f["gan"], f["code"]))
    volume.sort(key=lambda f: (-f["gan"], -f["gmax"], -f["score"], f["code"]))
    selected_core = core[0] if core and not noise_blocked else None
    selected_volume = volume[0] if volume and not noise_blocked and selected_core is None else None

    # Always display at least one A1 candidate and its earliest milestone.
    display: list[tuple[str, dict[str, Any]]] = []
    if selected_core:
        display.append(("CORE", selected_core))
    elif selected_volume:
        display.append(("VOLUME", selected_volume))
        # Add the strongest Core watch if different, to expose its earliest date.
        near_core = min(
            features.values(),
            key=lambda f: (
                0 if f["gmax"] > 0 else 1,
                max(21 - f["gan"], 0),
                abs(f["score"] - 0.90),
                -f["gmax"],
                f["code"],
            ),
        )
        if near_core["code"] != selected_volume["code"]:
            display.append(("CORE_WATCH", near_core))
    else:
        near_volume = min(
            features.values(),
            key=lambda f: (
                0 if f["gmax"] > 0 else 1,
                max(12 - f["gan"], 0),
                abs(f["score"] - 0.70),
                f["presence5"],
                -f["gan"],
                f["code"],
            ),
        )
        display.append(("VOLUME_WATCH", near_volume))

    candidates: list[dict[str, Any]] = []
    for rank, (kind, feature) in enumerate(display, start=1):
        tier = "CORE" if "CORE" in kind else "VOLUME"
        earliest, condition, milestone_type = earliest_a1_date(
            feature, tier, target_date, noise_blocked
        )
        gate = (selected_core is feature) or (selected_volume is feature)
        reason = (
            f"Gan {feature['gan']}; Gmax {feature['gmax']}; Score {feature['score']:.3f}; "
            f"Presence5 {feature['presence5']}; MaxFreq5 {feature['maxfreq5']}."
        )
        if noise_blocked:
            reason += f" Phanh nhiễu: repeat2={repeat2_count}, maxfreq={max_frequency}."
        candidates.append(
            {
                "code": feature["code"],
                "rank": rank,
                "gate": gate,
                "status": (
                    "CORE PASS" if selected_core is feature else "VOLUME PASS" if selected_volume is feature else kind
                ),
                "gan": feature["gan"],
                "gmax": feature["gmax"],
                "score": round(feature["score"], 6),
                "hot21": feature["hot21"],
                "reason": reason,
                "earliest_eligible_date": iso_day(earliest),
                "earliest_condition": condition,
                "milestone_type": milestone_type,
            }
        )
    return {
        "noise_blocked": noise_blocked,
        "core": selected_core,
        "volume": selected_volume,
        "candidates": candidates,
    }


def build_x3(
    features: dict[str, dict[str, Any]], repeat2_count: int, max_frequency: int, target_date: date
) -> dict[str, Any]:
    eligible_pool = [feature for feature in features.values() if not feature["in_latest"]]
    eligible_pool.sort(
        key=lambda f: (-f["x3_score"], -f["hot21"], -f["gan"], f["code"])
    )
    basket = eligible_pool[:3]
    hot_sum = sum(int(item["hot21"]) for item in basket)
    hot_max = max((int(item["hot21"]) for item in basket), default=0)
    gate = (
        len(basket) == 3
        and repeat2_count <= 4
        and max_frequency <= 2
        and 32 <= hot_sum <= 34
        and 11 <= hot_max <= 14
    )
    earliest = target_date if gate else target_date + timedelta(days=1)
    condition = (
        "Đủ X3 Growth khi rổ top3 giữ HOT21 tổng 32–34, max 11–14, repeat2≤4, "
        "maxfreq≤2 và không dùng fallback. Rổ phải được xếp lại sau mỗi kỳ khóa."
    )
    candidates = []
    for rank, feature in enumerate(basket, start=1):
        candidates.append(
            {
                "code": feature["code"],
                "rank": rank,
                "gate": gate,
                "status": f"RANK {rank}" if not gate else "X3 PASS",
                "gan": feature["gan"],
                "hot21": feature["hot21"],
                "reason": f"X3Score {feature['x3_score']:.3f}",
                "earliest_eligible_date": iso_day(earliest),
                "earliest_condition": condition,
                "milestone_type": "ELIGIBLE_NOW" if gate else "DYNAMIC_LOWER_BOUND",
            }
        )
    return {
        "basket": basket,
        "gate": gate,
        "hot_sum": hot_sum,
        "hot_max": hot_max,
        "candidates": candidates,
        "earliest": earliest,
        "condition": condition,
    }


def inverse_pairs() -> list[tuple[str, str, str]]:
    pairs: list[tuple[str, str, str]] = []
    for number in range(100):
        code = f"{number:02d}"
        reverse = code[::-1]
        if code == reverse or int(code) < int(reverse):
            continue
        pairs.append((f"{code}-{reverse}", code, reverse))
    return pairs


def x2_pair_stats(
    features: dict[str, dict[str, Any]]
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for pair, main, cover in inverse_pairs():
        fm, fc = features[main], features[cover]
        primary = min(fm["h60"], fc["h60"]) / 60 + min(fm["h90"], fc["h90"]) / 90
        secondary = (fm["h60"] + fc["h60"]) / 60 + (fm["h90"] + fc["h90"]) / 90
        dgan = abs(fm["gan"] - fc["gan"])
        h5_pair = fm["h5"] + fc["h5"]
        min_h21 = min(fm["h21"], fc["h21"])
        core_gate = (
            1 <= fm["gan"] <= 5
            and 1 <= fc["gan"] <= 5
            and dgan <= 1
            and h5_pair <= 3
            and min_h21 >= 5
        )
        balanced_gate = (
            2 <= fm["gan"] <= 7
            and 2 <= fc["gan"] <= 7
            and dgan <= 2
            and h5_pair <= 7
            and min_h21 >= 5
        )
        tv21 = min_h21 + (fm["h21"] + fc["h21"]) / 100 + primary / 1000 + secondary / 10000
        rank_score = primary + secondary / 1000 + min_h21 / 100000
        rows.append(
            {
                "pair": pair,
                "main": main,
                "cover": cover,
                "gan_main": fm["gan"],
                "gan_cover": fc["gan"],
                "dgan": dgan,
                "h5_pair": h5_pair,
                "h21_main": fm["h21"],
                "h21_cover": fc["h21"],
                "min_h21": min_h21,
                "h60_main": fm["h60"],
                "h60_cover": fc["h60"],
                "h90_main": fm["h90"],
                "h90_cover": fc["h90"],
                "primary": primary,
                "secondary": secondary,
                "core_gate": core_gate,
                "balanced_gate": balanced_gate,
                "tv21": tv21,
                "rank_score": rank_score,
                "mask5_main": fm["mask5"],
                "mask5_cover": fc["mask5"],
                "mask21_main": fm["mask21"],
                "mask21_cover": fc["mask21"],
            }
        )
    return rows


def x2_gate_state(state: tuple[int, int, int, int, int, int], tier: str) -> bool:
    gm, gc, m5m, m5c, m21m, m21c = state
    h5_pair = m5m.bit_count() + m5c.bit_count()
    min_h21 = min(m21m.bit_count(), m21c.bit_count())
    dgan = abs(gm - gc)
    if tier == "CORE":
        return 1 <= gm <= 5 and 1 <= gc <= 5 and dgan <= 1 and h5_pair <= 3 and min_h21 >= 5
    return 2 <= gm <= 7 and 2 <= gc <= 7 and dgan <= 2 and h5_pair <= 7 and min_h21 >= 5


def earliest_x2_date(row: dict[str, Any], tier: str, target_date: date, horizon: int = 7) -> tuple[date, str, str]:
    initial = (
        int(row["gan_main"]),
        int(row["gan_cover"]),
        int(row["mask5_main"]),
        int(row["mask5_cover"]),
        int(row["mask21_main"]),
        int(row["mask21_cover"]),
    )
    states = {initial}
    mask5_limit = (1 << 5) - 1
    mask21_limit = (1 << 21) - 1
    for offset in range(horizon + 1):
        if any(x2_gate_state(state, tier) for state in states):
            return (
                target_date + timedelta(days=offset),
                f"Mốc lạc quan sớm nhất để cặp rank1 đạt {tier}; cần đúng chuỗi hit/miss và toàn bộ gate Gan/ΔGan/H5/H21. Tính lại sau mỗi kỳ khóa.",
                "ELIGIBLE_NOW" if offset == 0 else "CONDITIONAL_LOWER_BOUND",
            )
        next_states: set[tuple[int, int, int, int, int, int]] = set()
        for gm, gc, m5m, m5c, m21m, m21c in states:
            for hit_m in (0, 1):
                for hit_c in (0, 1):
                    next_states.add(
                        (
                            0 if hit_m else gm + 1,
                            0 if hit_c else gc + 1,
                            ((m5m << 1) | hit_m) & mask5_limit,
                            ((m5c << 1) | hit_c) & mask5_limit,
                            ((m21m << 1) | hit_m) & mask21_limit,
                            ((m21c << 1) | hit_c) & mask21_limit,
                        )
                    )
        states = next_states
    return (
        target_date + timedelta(days=1),
        f"Chưa có đường vào {tier} trong {horizon} kỳ mô phỏng; đây là mốc rà lại sớm nhất sau kỳ khóa tiếp theo, không phải cam kết vào lệnh.",
        "DYNAMIC_RECALCULATION",
    )


def x2_brake_active(brake: dict[str, Any]) -> tuple[bool, str]:
    status = str(brake.get("status", "")).upper()
    remaining = int(brake.get("brake_remaining", 0) or 0)
    loss_streak = int(brake.get("loss_streak", 0) or 0)
    drawdown = int(brake.get("drawdown_vnd", 0) or 0)
    active = remaining > 0 or loss_streak >= 5 or drawdown <= -3_450_000 or "SUSPEND" in status
    reason = (
        f"Pilot={status or 'ACTIVE'}; BrakeRemaining={remaining}; LossStreak={loss_streak}; DD={signed_vnd(drawdown)}."
    )
    return active, reason


def build_x2(
    features: dict[str, dict[str, Any]], brake: dict[str, Any], target_date: date
) -> dict[str, Any]:
    rows = x2_pair_stats(features)
    core_rank = sorted(rows, key=lambda r: (-r["tv21"], r["pair"]))[0]
    balanced_rank = sorted(rows, key=lambda r: (-r["rank_score"], r["pair"]))[0]
    brake_active, brake_reason = x2_brake_active(brake)
    if core_rank["core_gate"]:
        selected_tier, selected = "CORE", core_rank
    elif balanced_rank["balanced_gate"]:
        selected_tier, selected = "BALANCED", balanced_rank
    else:
        selected_tier, selected = None, None
    pass_gate = selected is not None and not brake_active

    displayed: list[tuple[str, dict[str, Any]]] = [("CORE", core_rank)]
    if balanced_rank["pair"] != core_rank["pair"]:
        displayed.append(("BALANCED", balanced_rank))
    candidates = []
    for rank, (tier, row) in enumerate(displayed, start=1):
        earliest, condition, milestone_type = earliest_x2_date(row, tier, target_date)
        raw_gate = bool(row["core_gate"] if tier == "CORE" else row["balanced_gate"])
        candidates.append(
            {
                "code": row["pair"],
                "rank": 1,
                "gate": raw_gate and not brake_active,
                "status": f"{tier} RANK1 · " + ("PASS" if raw_gate and not brake_active else "FAIL"),
                "reason": (
                    f"Gan {row['gan_main']}/{row['gan_cover']}; ΔGan {row['dgan']}; "
                    f"H5pair {row['h5_pair']}; minH21 {row['min_h21']}; "
                    f"TV21 {row['tv21']:.4f}; B {row['primary']:.4f}. {brake_reason}"
                ),
                "earliest_eligible_date": iso_day(earliest),
                "earliest_condition": condition,
                "milestone_type": milestone_type,
            }
        )
    return {
        "core_rank": core_rank,
        "balanced_rank": balanced_rank,
        "selected_tier": selected_tier,
        "selected": selected,
        "pass_gate": pass_gate,
        "brake_active": brake_active,
        "brake_reason": brake_reason,
        "candidates": candidates,
    }


def scenario_text(points: int, code_count: int) -> str:
    capital = points * LO_COST_PER_POINT * code_count
    scenarios = []
    for hits in range(0, code_count + 1):
        pnl = hits * points * LO_PAYOUT_PER_HIT_POINT - capital
        scenarios.append(f"{hits} nháy {signed_vnd(pnl)}")
    return " · ".join(scenarios)


def make_plan(
    doc: dict[str, Any], history: list[tuple[date, list[str]]], history_source: str, brake: dict[str, Any]
) -> tuple[dict[str, Any], dict[str, Any]]:
    locked_date, latest_codes = history[-1]
    target_date = locked_date + timedelta(days=1)
    latest_counts = Counter(latest_codes)
    repeat2_count = sum(1 for count in latest_counts.values() if count >= 2)
    max_frequency = max(latest_counts.values())
    unique = len(latest_counts)
    features = build_features(history)
    a1 = a1_candidates(features, repeat2_count, max_frequency, target_date)
    x3 = build_x3(features, repeat2_count, max_frequency, target_date)
    x2 = build_x2(features, brake, target_date)

    decision = "A0"
    selection: list[str] = []
    points_per_code = 0
    method_label = "A0"
    reason = "Không có phương pháp đạt toàn bộ gate."
    selected_method_id = "A0"
    if a1["core"] is not None:
        decision = "A1_CORE100"
        selection = [a1["core"]["code"]]
        points_per_code = A1_CORE_POINTS
        method_label = "A1 Core Strict"
        reason = "A1 Core đứng ưu tiên 1 và đạt toàn bộ Gan/Score/cooldown/noise gate."
        selected_method_id = "A1_CORE"
    elif a1["volume"] is not None:
        decision = "A1_VOLUME50"
        selection = [a1["volume"]["code"]]
        points_per_code = A1_VOLUME_POINTS
        method_label = "A1 Volume Balanced"
        reason = "Không có Core; A1 Volume đạt toàn bộ gate và đứng trước X3/X2."
        selected_method_id = "A1_VOLUME"
    elif x3["gate"]:
        decision = "X3_GROWTH50"
        selection = [item["code"] for item in x3["basket"]]
        points_per_code = X3_POINTS
        method_label = "MB X3 Growth32-34"
        reason = "A1 A0; X3 Growth top3 đạt HOT21 sum/max và noise gate."
        selected_method_id = "X3_GROWTH"
    elif x2["pass_gate"] and x2["selected"] is not None:
        decision = "X2_RESCUE15"
        selection = [x2["selected"]["main"], x2["selected"]["cover"]]
        points_per_code = X2_POINTS
        method_label = "MB X2 Rescue"
        reason = "A1 và X3 A0; cặp X2 rank1 đạt gate và không bị phanh pilot."
        selected_method_id = "X2_RESCUE"

    capital = points_per_code * LO_COST_PER_POINT * len(selection)
    pending = decision != "A0"

    # Preserve settlement/ledger fields but replace the prospective decision layer.
    out = copy.deepcopy(doc)
    if out.get("settlement"):
        out["last_settlement"] = copy.deepcopy(out["settlement"])
    out["schema_version"] = "MB_DAILY_WEB_V7_AUTO_PLAN"
    out["target_date"] = iso_day(target_date)
    out["valid_until"] = datetime.combine(
        target_date, datetime.max.time().replace(microsecond=0), tzinfo=VN
    ).isoformat()
    out["data_snapshot_id"] = f"MBH27_{locked_date.strftime('%Y%m%d')}_{json_hash(latest_codes)[:12].upper()}"
    out["stake_rule"] = {
        "a1_core_points_per_code": A1_CORE_POINTS,
        "a1_volume_points_per_code": A1_VOLUME_POINTS,
        "x3_points_per_code": X3_POINTS,
        "x2_points_per_code": X2_POINTS,
        "lo_cost_per_point_vnd": LO_COST_PER_POINT,
        "lo_payout_per_hit_point_vnd": LO_PAYOUT_PER_HIT_POINT,
        "controller_order": CONTROLLER_ORDER,
        "note": "Tối đa một lệnh tiền thật; chưa xác nhận thì không ghi P/L.",
    }
    out["milestone_policy"] = {
        "status": "ACTIVE_MANDATORY_ABSOLUTE",
        "source": "AUTO_EARLIEST_ENTRY_ENGINE_V1",
        "rule": "Mọi ứng viên A1/X2/X3 hiển thị phải có earliest_eligible_date, điều kiện và loại mốc.",
        "recalculate_after_each_locked_draw": True,
        "disclosure": "Mốc sớm nhất là lower bound có điều kiện, không phải cam kết. Mọi mốc được tính lại sau mỗi kỳ khóa đủ 27/27.",
    }
    out["top_signal_policy"] = {
        "mode": "PROSPECTIVE_AUTO_PLAN",
        "controller": ">".join(CONTROLLER_ORDER),
        "system_signals_preserved_in_groups": True,
        "pnl_rule": "Chưa có xác nhận đánh thật thì không cộng lãi/lỗ.",
    }
    out["portfolio"] = {
        "decision": decision,
        "tier": ("A0" if decision == "A0" else f"{method_label} · CHỜ XÁC NHẬN"),
        "selection": "-".join(selection) if selection else "A0",
        "title": "A0 — KHÔNG CÓ LỆNH" if decision == "A0" else f"{method_label.upper()} {'-'.join(selection)} — {points_per_code} ĐIỂM/SỐ",
        "points": points_per_code * len(selection),
        "capital_vnd": capital,
        "payout_vnd": 0,
        "pnl_vnd": 0,
        "reason": reason,
        "pnl_status": "NOT_INCLUDED_UNTIL_CONFIRMED",
    }
    if pending:
        out["pending_order"] = {
            "status": "SYSTEM_SIGNAL_NOT_YET_CONFIRMED",
            "date": iso_day(target_date),
            "method_id": selected_method_id,
            "selection": selection,
            "points_per_code": points_per_code,
            "capital_vnd": capital,
            "pnl_included": False,
            "note": "Chỉ chuyển thành actual_order khi người dùng xác nhận trước quay.",
        }
    else:
        out.pop("pending_order", None)

    method_numbers = [
        {
            "code": code,
            "points": points_per_code,
            "capital_vnd": points_per_code * LO_COST_PER_POINT,
            "role": "Tín hiệu hệ thống; chưa xác nhận",
        }
        for code in selection
    ]
    out["top_signals"] = {
        "title": f"KẾ HOẠCH TỰ ĐỘNG {target_date.strftime('%d/%m/%Y')}",
        "subtitle": "A0" if not pending else f"{method_label} · {'-'.join(selection)} ×{points_per_code} điểm",
        "total_methods": 0 if not pending else 1,
        "total_numbers": f"{len(selection)} mã" if selection else "0 mã",
        "total_points": f"{points_per_code * len(selection)} điểm",
        "total_capital_vnd": capital,
        "note": (
            f"{reason} Kịch bản: {scenario_text(points_per_code, len(selection))}."
            if pending
            else "Không có lệnh; tất cả ứng viên giữ Watch/Shadow với mốc sớm nhất."
        ),
        "methods": []
        if not pending
        else [
            {
                "id": selected_method_id,
                "label": method_label,
                "method": decision,
                "status": "PASS_PENDING_CONFIRMATION",
                "points_per_code": points_per_code,
                "code_count": len(selection),
                "capital_vnd": capital,
                "numbers": method_numbers,
            }
        ],
    }

    # A1 group.
    a1_selected = a1["core"] or a1["volume"]
    a1_status = "PASS_CORE" if a1["core"] else "PASS_VOLUME" if a1["volume"] else "A0_WATCH"
    replace_group(
        out,
        {
            "id": "A1",
            "label": "MB A1",
            "status": a1_status,
            "role": "PRODUCTION MAIN",
            "method": "Dual Core Volume Balance — Core100/Volume50",
            "layer": "Core → Volume",
            "selected_numbers": [a1_selected["code"]] if a1_selected else [a1["candidates"][0]["code"]],
            "points": points_per_code if decision.startswith("A1_") else 0,
            "capital_vnd": capital if decision.startswith("A1_") else 0,
            "summary": (
                f"Đạt {a1_status}; mã {a1_selected['code']}." if a1_selected else "Không có A1 đạt; hiển thị Watch và mốc sớm nhất."
            ),
            "reason": (
                f"repeat2={repeat2_count}; maxfreq={max_frequency}; hard brake={'ON' if a1['noise_blocked'] else 'OFF'}."
            ),
            "candidates": a1["candidates"],
        },
    )

    # X3 group.
    replace_group(
        out,
        {
            "id": "X3",
            "label": "MB X3",
            "status": "PASS_GROWTH" if x3["gate"] else "A0_SHADOW",
            "role": "PRODUCTION SUPPORT",
            "method": "MB X3 Growth32-34",
            "layer": "Generator top3",
            "selected_numbers": [item["code"] for item in x3["basket"]],
            "points": points_per_code * len(selection) if decision.startswith("X3_") else 0,
            "capital_vnd": capital if decision.startswith("X3_") else 0,
            "summary": f"HOT21 sum={x3['hot_sum']}; max={x3['hot_max']}; {'PASS' if x3['gate'] else 'FAIL'}.",
            "reason": "Không fallback; rổ được xếp lại sau mỗi kỳ khóa.",
            "candidates": x3["candidates"],
            "earliest_basket_date": iso_day(x3["earliest"]),
            "earliest_basket_condition": x3["condition"],
        },
    )

    # X2 group. No performance/backtest metrics are ever emitted.
    x2_selected_pairs = [x2["core_rank"]["pair"]]
    if x2["balanced_rank"]["pair"] != x2["core_rank"]["pair"]:
        x2_selected_pairs.append(x2["balanced_rank"]["pair"])
    replace_group(
        out,
        {
            "id": "X2",
            "label": "MB X2",
            "status": "PASS_RESCUE" if x2["pass_gate"] else "A0_SHADOW",
            "role": "GUARDED RESCUE",
            "method": "MB X2 Fusion A0 Rescue",
            "layer": "Core rank1 → Balanced rank1",
            "selected_numbers": x2_selected_pairs,
            "points": points_per_code * len(selection) if decision.startswith("X2_") else 0,
            "capital_vnd": capital if decision.startswith("X2_") else 0,
            "summary": (
                f"Cặp đạt: {x2['selected']['pair']} ({x2['selected_tier']})."
                if x2["pass_gate"] and x2["selected"]
                else "Core rank1/Balanced rank1 chưa tạo lệnh X2."
            ),
            "reason": x2["brake_reason"],
            "candidates": x2["candidates"],
            "display_policy": {
                "hide_fields": [
                    "historical_profile",
                    "quality_benchmark",
                    "orders",
                    "wins",
                    "losses",
                    "win_rate",
                    "oos_win_rate",
                    "pnl_vnd",
                    "pnl_units",
                    "max_drawdown_vnd",
                    "max_drawdown_units",
                ],
                "keep_fields": [],
                "scope": "WEBSITE_X2_NO_PERFORMANCE_METRICS_PERMANENT",
            },
        },
    )

    # Keep Xiên and Đề as separate, never auto-created real orders.
    xien = find_group(out, "XIEN") or {"id": "XIEN", "label": "Xiên 2"}
    xien.update(
        {
            "status": "A0_SHADOW",
            "role": "SỔ RIÊNG",
            "method": "Không tự mở từ A1/X2/X3",
            "layer": "Chờ xác nhận riêng",
            "selected_numbers": [],
            "points": 0,
            "capital_vnd": 0,
            "summary": "Không có lệnh Xiên tự động.",
            "reason": "Chỉ mở khi người dùng xác nhận riêng trước quay.",
        }
    )
    replace_group(out, xien)
    de = find_group(out, "DE") or {"id": "DE", "label": "Đề đầu/đuôi"}
    de.update(
        {
            "status": "A0_VALIDATION",
            "role": "VALIDATION SHADOW",
            "selected_numbers": [],
            "points": 0,
            "capital_vnd": 0,
            "summary": "Không có lệnh thật tự động.",
            "reason": "Chỉ theo dõi theo gate riêng.",
        }
    )
    replace_group(out, de)

    q = out.setdefault("pnl_summary", {})
    q["today_pending_capital_vnd"] = capital
    q["today_pending_order"] = (
        "A0 — không có lệnh" if not pending else f"{method_label} {'-'.join(selection)} ×{points_per_code}; chờ xác nhận"
    )
    q["today_included"] = False

    # Hash only deterministic review content, not timestamps.
    plan_core = {
        "target_date": iso_day(target_date),
        "locked_through": iso_day(locked_date),
        "data_hash": json_hash(latest_codes),
        "controller": CONTROLLER_ORDER,
        "portfolio": out["portfolio"],
        "top_signals": out["top_signals"],
        "groups": [find_group(out, group_id) for group_id in MILESTONE_GROUPS],
        "milestone_policy": out["milestone_policy"],
    }
    plan_hash = json_hash(plan_core)
    run_id = f"RPT_MB_{target_date.strftime('%Y%m%d')}_AUTO_{plan_hash[:12].upper()}"
    out["report_run_id"] = run_id
    out["source_run_id"] = f"LOCK_{locked_date.strftime('%Y%m%d')}_{json_hash(latest_codes)[:10].upper()}"
    out["config_id"] = "MB_AUTO_CONTROLLER_CORE100_VOLUME50_X3_X2_V1"

    generated = now_vn().isoformat(timespec="seconds")
    out["generated_at"] = generated
    out["automation"] = {
        "status": "AUTO_PLAN_READY_FOR_WEB",
        "pipeline_version": "MB_DAILY_PIPELINE_V1",
        "source": f"GOOGLE_SHEET_XLSX:{history_source}",
        "last_updated_at": generated,
        "locked_through": iso_day(locked_date),
        "target_date": iso_day(target_date),
        "settlement_checked": True,
        "signal_review_complete": True,
        "milestones_complete": True,
        "plan_hash": plan_hash,
        "website_refresh_seconds": 120,
    }

    plan_snapshot = {
        "schema_version": "MB_DAILY_PLAN_SNAPSHOT_V1",
        "report_run_id": run_id,
        "plan_hash": plan_hash,
        "generated_at": generated,
        "locked_through": iso_day(locked_date),
        "target_date": iso_day(target_date),
        "data_snapshot": out.get("data"),
        "controller_order": CONTROLLER_ORDER,
        "portfolio": out["portfolio"],
        "top_signals": out["top_signals"],
        "groups": [find_group(out, group_id) for group_id in ("A1", "X2", "X3", "XIEN", "DE")],
        "milestone_policy": out["milestone_policy"],
        "validation": {
            "data_27_ok": len(latest_codes) == 27,
            "milestones_complete": True,
            "x2_performance_metrics_absent": True,
            "pnl_not_included_without_confirmation": True,
        },
    }
    validate_plan(out)
    return out, plan_snapshot


def validate_plan(doc: dict[str, Any]) -> None:
    if (doc.get("milestone_policy") or {}).get("status") != "ACTIVE_MANDATORY_ABSOLUTE":
        raise RuntimeError("Milestone policy chưa khóa ở ACTIVE_MANDATORY_ABSOLUTE")
    for group_id in MILESTONE_GROUPS:
        group = find_group(doc, group_id)
        if not group:
            raise RuntimeError(f"Thiếu group {group_id}")
        candidates = group.get("candidates") or []
        if not candidates:
            raise RuntimeError(f"{group_id} không có ứng viên để hiển thị mốc")
        for candidate in candidates:
            for key in ("earliest_eligible_date", "earliest_condition", "milestone_type"):
                if not candidate.get(key):
                    raise RuntimeError(f"{group_id}/{candidate.get('code')} thiếu {key}")
        if group_id == "X3" and not group.get("earliest_basket_date"):
            raise RuntimeError("X3 thiếu earliest_basket_date")
    x2 = find_group(doc, "X2") or {}
    forbidden = {
        "historical_profile",
        "quality_benchmark",
        "orders",
        "wins",
        "losses",
        "win_rate",
        "oos_win_rate",
        "pnl_vnd",
        "pnl_units",
        "max_drawdown_vnd",
        "max_drawdown_units",
    }
    leaked = forbidden.intersection(x2)
    if leaked:
        raise RuntimeError(f"X2 làm rò chỉ số hiệu suất: {sorted(leaked)}")
    pending = doc.get("pending_order")
    if pending and pending.get("pnl_included"):
        raise RuntimeError("pending_order không được phép cộng P/L")


def persist_plan(doc: dict[str, Any], snapshot: dict[str, Any]) -> bool:
    plan_hash = snapshot["plan_hash"]
    current = load_json(DATA_FILE, {})
    current_hash = (current.get("automation") or {}).get("plan_hash")
    same_plan = (
        current_hash == plan_hash
        and current.get("target_date") == snapshot["target_date"]
        and (current.get("automation") or {}).get("milestones_complete") is True
    )
    changed = False
    if not same_plan:
        changed |= write_json_if_changed(DATA_FILE, doc)

    PLANS_DIR.mkdir(parents=True, exist_ok=True)
    plan_path = PLANS_DIR / f"{snapshot['target_date']}.json"
    # Dated plan snapshots are replaced only when their deterministic hash changes.
    existing_snapshot = load_json(plan_path, {})
    if existing_snapshot.get("plan_hash") != plan_hash:
        changed |= write_json_if_changed(plan_path, snapshot)

    ledger = load_json(
        PLAN_LEDGER_FILE,
        {"schema_version": "MB_DAILY_PLAN_LEDGER_V1", "latest_target_date": None, "plans": {}},
    )
    entry = {
        "path": f"data/plans/{snapshot['target_date']}.json",
        "report_run_id": snapshot["report_run_id"],
        "plan_hash": plan_hash,
        "locked_through": snapshot["locked_through"],
        "decision": snapshot["portfolio"]["decision"],
        "selection": snapshot["portfolio"]["selection"],
        "milestones_complete": True,
    }
    if ledger.get("plans", {}).get(snapshot["target_date"]) != entry or ledger.get("latest_target_date") != snapshot["target_date"]:
        ledger.setdefault("plans", {})[snapshot["target_date"]] = entry
        ledger["latest_target_date"] = snapshot["target_date"]
        ledger["updated_at"] = snapshot["generated_at"]
        changed |= write_json_if_changed(PLAN_LEDGER_FILE, ledger)

    state = {
        "schema_version": "MB_AUTOMATION_STATE_V1",
        "status": "OK",
        "pipeline_version": "MB_DAILY_PIPELINE_V1",
        "last_locked_through": snapshot["locked_through"],
        "next_target_date": snapshot["target_date"],
        "report_run_id": snapshot["report_run_id"],
        "plan_hash": plan_hash,
        "milestones_complete": True,
        "website_payload": "data/current.json",
        "plan_snapshot": f"data/plans/{snapshot['target_date']}.json",
        "updated_at": snapshot["generated_at"],
    }
    existing_state = load_json(AUTOMATION_STATE_FILE, {})
    if existing_state.get("plan_hash") != plan_hash or existing_state.get("status") != "OK":
        changed |= write_json_if_changed(AUTOMATION_STATE_FILE, state)
    return changed


def data_fail(reason: str) -> bool:
    doc = load_json(DATA_FILE, {})
    data = doc.get("data") or {}
    locked = parse_date(data.get("locked_through")) or now_vn().date()
    target = locked + timedelta(days=1)
    normalized = re.sub(r"\s+", " ", str(reason or "Nguồn dữ liệu không hợp lệ")).strip()[:500]
    generated = now_vn().isoformat(timespec="seconds")
    fail_hash = json_hash({"locked": iso_day(locked), "reason": normalized})
    doc["schema_version"] = "MB_DAILY_WEB_V7_AUTO_PLAN"
    doc["target_date"] = iso_day(target)
    doc["report_run_id"] = f"RPT_MB_{target.strftime('%Y%m%d')}_DATA_FAIL_{fail_hash[:10].upper()}"
    doc["portfolio"] = {
        "decision": "A0_DATA_FAIL",
        "tier": "A0 — DỮ LIỆU CHƯA KHÓA",
        "selection": "A0",
        "title": "A0 — CHỜ DỮ LIỆU 27/27",
        "points": 0,
        "capital_vnd": 0,
        "payout_vnd": 0,
        "pnl_vnd": 0,
        "reason": normalized,
        "pnl_status": "NOT_INCLUDED",
    }
    doc.pop("pending_order", None)
    doc["top_signals"] = {
        "title": "DỮ LIỆU KHÔNG ĐỦ ĐỂ RÀ SOÁT",
        "subtitle": "A0",
        "total_methods": 0,
        "total_numbers": "0 mã",
        "total_points": "0 điểm",
        "total_capital_vnd": 0,
        "note": "Không suy đoán. Pipeline sẽ thử lại theo lịch 19:35/19:55 hoặc lần 14:00 kế tiếp.",
        "methods": [],
    }
    doc["milestone_policy"] = {
        "status": "ACTIVE_MANDATORY_ABSOLUTE",
        "source": "AUTO_EARLIEST_ENTRY_ENGINE_V1",
        "rule": "Không có mốc vào lệnh khi dữ liệu fail; mốc sớm nhất là lần khóa đủ 27/27 kế tiếp.",
        "recalculate_after_each_locked_draw": True,
        "disclosure": "A0 bắt buộc cho đến khi nguồn đủ 27/27 và đối chiếu thành công.",
    }
    for group_id, label in (("A1", "MB A1"), ("X2", "MB X2"), ("X3", "MB X3")):
        group = {
            "id": group_id,
            "label": label,
            "status": "A0_DATA_FAIL",
            "role": "WAITING_FOR_LOCK",
            "method": "Không tính khi dữ liệu chưa khóa",
            "layer": "DATA GATE",
            "selected_numbers": [],
            "points": 0,
            "capital_vnd": 0,
            "summary": "Không có ứng viên hợp lệ.",
            "reason": normalized,
            "candidates": [
                {
                    "code": "—",
                    "rank": 1,
                    "gate": False,
                    "status": "DATA FAIL",
                    "reason": "Chờ đủ 27/27 và đối chiếu nguồn.",
                    "earliest_eligible_date": iso_day(target),
                    "earliest_condition": "Chỉ được rà lại sau khi dữ liệu khóa đủ 27/27; không phải mốc chắc chắn vào lệnh.",
                    "milestone_type": "NEXT_DATA_LOCK",
                }
            ],
        }
        if group_id == "X3":
            group["earliest_basket_date"] = iso_day(target)
            group["earliest_basket_condition"] = "Chờ kỳ khóa đủ 27/27 để xếp lại rổ."
        if group_id == "X2":
            group["display_policy"] = {
                "hide_fields": ["historical_profile", "quality_benchmark", "orders", "wins", "losses", "win_rate", "pnl_vnd", "max_drawdown_vnd"],
                "keep_fields": [],
                "scope": "WEBSITE_X2_NO_PERFORMANCE_METRICS_PERMANENT",
            }
        replace_group(doc, group)
    doc["generated_at"] = generated
    doc["automation"] = {
        "status": "DATA_FAIL_A0",
        "pipeline_version": "MB_DAILY_PIPELINE_V1",
        "reason": normalized,
        "last_updated_at": generated,
        "locked_through": iso_day(locked),
        "target_date": iso_day(target),
        "signal_review_complete": False,
        "milestones_complete": True,
        "plan_hash": fail_hash,
        "website_refresh_seconds": 120,
    }
    validate_plan(doc)
    changed = write_json_if_changed(DATA_FILE, doc)
    state = {
        "schema_version": "MB_AUTOMATION_STATE_V1",
        "status": "DATA_FAIL_A0",
        "reason": normalized,
        "last_locked_through": iso_day(locked),
        "next_target_date": iso_day(target),
        "milestones_complete": True,
        "updated_at": generated,
    }
    changed |= write_json_if_changed(AUTOMATION_STATE_FILE, state)
    return changed


def self_test() -> None:
    target = date(2026, 7, 12)
    # Balanced pair with Gan 0/1 can reach 2/3 after two misses.
    row = {
        "gan_main": 0,
        "gan_cover": 1,
        "mask5_main": 0,
        "mask5_cover": 0,
        "mask21_main": (1 << 6) - 1,
        "mask21_cover": (1 << 6) - 1,
    }
    when, _, _ = earliest_x2_date(row, "BALANCED", target)
    assert when == date(2026, 7, 14), when
    # A1 eligible now when all gates are clean.
    feature = {"gan": 14, "gmax": 20, "last5_counts": (0, 0, 0, 0, 0)}
    when, _, milestone = earliest_a1_date(feature, "VOLUME", target, False)
    assert when == target and milestone == "ELIGIBLE_NOW"
    assert len(inverse_pairs()) == 45
    print("SELF_TEST_OK")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--data-fail", help="Write a safe A0 payload instead of planning")
    parser.add_argument("--self-test", action="store_true")
    args = parser.parse_args()
    if args.self_test:
        self_test()
        return
    if args.data_fail:
        changed = data_fail(args.data_fail)
        print(f"DATA_FAIL_A0 written={changed}")
        return
    if not DATA_FILE.exists():
        raise RuntimeError(f"Thiếu {DATA_FILE}")
    doc = load_json(DATA_FILE, {})
    xlsx_path, temp_handle = obtain_xlsx()
    try:
        history, source, brake = load_history_and_brake(xlsx_path, doc)
        planned, snapshot = make_plan(doc, history, source, brake)
        changed = persist_plan(planned, snapshot)
        print(
            f"PLAN_OK target={snapshot['target_date']} decision={snapshot['portfolio']['decision']} "
            f"selection={snapshot['portfolio']['selection']} milestones=OK changed={changed}"
        )
    finally:
        if temp_handle is not None:
            temp_handle.close()


if __name__ == "__main__":
    main()
