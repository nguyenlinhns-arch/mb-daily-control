#!/usr/bin/env python3
"""Áp dụng điều chỉnh cùng ngày đã được xác nhận trong Validation_Log.

Phanh an toàn:
- Chỉ xét kỳ mới nhất có đủ 27 mã.
- Tối thiểu hai tab nguồn phải cho cùng một bộ mã.
- Validation_Log phải có trạng thái OK_LOCKED_CROSSCHECKED_CORRECTION
  và chứa đúng chuỗi 27 mã của kỳ đó.
- Chỉ sửa snapshot website đang khóa CROSSCHECKED cùng ngày.
- Quyết toán lại bằng settlement ledger hiện có nên chỉ cộng phần chênh lệch.
"""
from __future__ import annotations

import copy
import os
import tempfile
import urllib.request
from datetime import date, datetime, time, timedelta
from typing import Any

from openpyxl import load_workbook

import sync_from_google_sheet as sync

SOURCE_XLSX_PATH = os.getenv('SOURCE_XLSX_PATH')
SOURCE_SHEETS = sync.SOURCE_SHEETS


def source_codes(sheet_name: str, row: tuple[Any, ...]) -> list[str | None]:
    if sheet_name == 'Raw_Results_IMPORT':
        if len(row) < 29 or str(row[1] or '').strip().upper() != 'XSMB':
            return []
        values = row[2:29]
    else:
        values = row[1:28]
    return [sync.code2(value) for value in values]


def authorised_correction(xlsx_path: str) -> tuple[date, list[str], list[str]] | None:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        per_sheet: list[tuple[date, list[str], str]] = []
        for name in SOURCE_SHEETS:
            if name not in wb.sheetnames:
                continue
            best: tuple[date, list[str], str] | None = None
            for row in wb[name].iter_rows(values_only=True):
                if not row:
                    continue
                draw_date = sync.as_date(row[0])
                codes = source_codes(name, row)
                if draw_date and len(codes) == 27 and all(codes):
                    clean = [str(code) for code in codes]
                    if best is None or draw_date > best[0]:
                        best = (draw_date, clean, name)
            if best:
                per_sheet.append(best)
        if not per_sheet:
            return None
        newest = max(item[0] for item in per_sheet)
        latest = [item for item in per_sheet if item[0] == newest]
        variants = {tuple(item[1]) for item in latest}
        if len(variants) > 1:
            names = ', '.join(item[2] for item in latest)
            raise RuntimeError(f'Lệch dữ liệu giữa các tab cùng ngày {newest}: {names}')
        if len(latest) < 2 or 'Validation_Log' not in wb.sheetnames:
            return None
        codes = latest[0][1]
        expected = ','.join(codes)
        authorised = False
        for row in wb['Validation_Log'].iter_rows(min_col=1, max_col=5, values_only=True):
            if not row or len(row) < 5:
                continue
            logged_date = sync.as_date(str(row[0] or '').strip()[:10])
            status = str(row[4] or '').strip().upper()
            actual = str(row[3] or '').replace(' ', '')
            if logged_date == newest and status == 'OK_LOCKED_CROSSCHECKED_CORRECTION' and expected in actual:
                authorised = True
                break
        if not authorised:
            return None
        return newest, codes, [item[2] for item in latest]
    finally:
        wb.close()


def apply_from_file(xlsx_path: str) -> None:
    correction = authorised_correction(xlsx_path)
    if correction is None:
        print('Không có điều chỉnh cùng ngày được cấp quyền; giữ phanh CROSSCHECKED.')
        return
    draw_date, codes, source_tabs = correction
    if draw_date > sync.now_vn().date():
        raise RuntimeError(f'Nguồn điều chỉnh có ngày tương lai {draw_date}; dừng tự động')

    doc = sync.load_json(sync.DATA_FILE, {})
    ledger = sync.load_json(sync.LEDGER_FILE, sync.default_ledger())
    existing = doc.get('data') or {}
    locked = sync.as_date(existing.get('locked_through'))
    current_codes = existing.get('latest_27_codes') or []
    current_status = str(existing.get('status', ''))

    if locked != draw_date:
        print(f'Điều chỉnh {draw_date} không trùng ngày website đang khóa {locked}; giao cho luồng đồng bộ chuẩn.')
        return
    if current_codes == codes:
        print(f'Website đã dùng đúng bộ 27 mã điều chỉnh ngày {draw_date}; không ghi lại.')
        return
    if 'CROSSCHECKED' not in current_status:
        print(f'Snapshot {draw_date} chưa ở trạng thái CROSSCHECKED; giao cho luồng đồng bộ chuẩn.')
        return

    before_doc = copy.deepcopy(doc)
    before_ledger = copy.deepcopy(ledger)
    snapshot = sync.build_data_snapshot(draw_date, codes, source_tabs)
    snapshot['status'] = 'LOCKED_CROSSCHECKED_CORRECTED'
    snapshot['source_label'] = f"Google Sheet XSMB · điều chỉnh đã đối chiếu {draw_date.strftime('%d/%m/%Y')} · đủ 27/27"
    doc['data'] = snapshot
    doc['target_date'] = draw_date.isoformat()
    doc['data_snapshot_id'] = f"CORRECTION_{draw_date.strftime('%Y%m%d')}_{sync.json_hash(codes)[:10].upper()}"
    doc['valid_until'] = datetime.combine(draw_date + timedelta(days=1), time(13, 59, 59), tzinfo=sync.VN).isoformat()

    settled = sync.settle_order(doc, ledger, draw_date, codes)
    changed_doc = doc != before_doc
    changed_ledger = ledger != before_ledger
    if not changed_doc and not changed_ledger:
        print(f'Điều chỉnh {draw_date} không làm thay đổi dữ liệu.')
        return

    timestamp = sync.now_vn().isoformat(timespec='seconds')
    if changed_doc:
        doc['generated_at'] = timestamp
        doc['automation'] = {
            'status': 'OK',
            'source': 'GOOGLE_SHEET_XLSX_CROSSCHECKED_CORRECTION',
            'last_updated_at': timestamp,
            'locked_through': draw_date.isoformat(),
            'settlement_checked': settled,
            'correction_authorised': True,
            'source_tabs': source_tabs,
        }
        sync.write_json(sync.DATA_FILE, doc)
    if changed_ledger:
        ledger['updated_at'] = timestamp
        sync.write_json(sync.LEDGER_FILE, ledger)
    print(f'Đã áp điều chỉnh được cấp quyền {draw_date}: settlement={settled}, current_json={changed_doc}, ledger={changed_ledger}.')


def main() -> None:
    if SOURCE_XLSX_PATH:
        apply_from_file(SOURCE_XLSX_PATH)
        return
    with tempfile.NamedTemporaryFile(suffix='.xlsx') as tmp:
        request = urllib.request.Request(sync.EXPORT_URL, headers={'User-Agent': 'MB-Daily-Control/2.1'})
        with urllib.request.urlopen(request, timeout=90) as response:
            tmp.write(response.read())
            tmp.flush()
        apply_from_file(tmp.name)


if __name__ == '__main__':
    main()
