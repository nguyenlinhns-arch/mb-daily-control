#!/usr/bin/env python3
"""Synchronize the public Lê Miền Bắc website snapshot from MB ALL Production.

Canonical daily source:
  XSMB_Source_2024_2026_MB_v1.3 / MAX2_Daily_Plan
Production method:
  MAX2_V1_R4268_P0072_HR60

The script deliberately excludes current Production codes from every public JSON.
Paid codes remain in the private payment workbook and are delivered only after
the owner approves the payment email.
"""
from __future__ import annotations

import argparse
import hashlib
import json
from collections import Counter
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

import build_public_methods as public_methods

ROOT = Path(__file__).resolve().parents[1]
VN = timezone(timedelta(hours=7))
PRODUCTION_METHOD = "MAX2_V1_R4268_P0072_HR60"
PRODUCTION_CONFIG = "MAX2_PRIMARY_V1_20260819"
SOURCE_SHEET_ID = "1iVAfqmS-TvP02U8FtKSM2nr_7Dsd7qi2qEGnWV6IK7w"
LEGACY_BRIDGE = {
    "2026-08-17": ["19", "91", "05", "50"],
    "2026-08-18": ["19", "91", "05", "50"],
}


def text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def iso_date(value: Any) -> str:
    raw = text(value)
    if not raw:
        return ""
    if len(raw) >= 10 and raw[4:5] == "-" and raw[7:8] == "-":
        return raw[:10]
    for fmt in ("%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(raw[:10], fmt).date().isoformat()
        except ValueError:
            pass
    return ""


def boolish(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return text(value).upper() in {"TRUE", "1", "YES", "Y"}


def number2(value: Any) -> str:
    if value is None:
        raise ValueError("empty code")
    if isinstance(value, bool):
        raise ValueError(f"invalid boolean code: {value!r}")
    if isinstance(value, (int, float)):
        numeric = float(value)
        if not numeric.is_integer():
            raise ValueError(f"non-integer code: {value!r}")
        number = int(numeric)
        if not 0 <= number <= 99:
            raise ValueError(f"code outside 00-99: {value!r}")
        return f"{number:02d}"
    raw = "".join(character for character in str(value).strip() if character.isdigit())
    if not raw:
        raise ValueError(f"empty or invalid code: {value!r}")
    return raw[-2:].zfill(2)


def rows_as_dicts(ws) -> list[dict[str, Any]]:
    values = list(ws.iter_rows(values_only=True))
    if not values:
        return []
    headers = [text(value) for value in values[0]]
    result = []
    for raw in values[1:]:
        if not any(value is not None and text(value) != "" for value in raw):
            continue
        result.append({headers[i]: raw[i] if i < len(raw) else None for i in range(len(headers)) if headers[i]})
    return result


def load_history(wb, lock_iso: str) -> list[tuple[date, list[str]]]:
    ws = wb["MB_History_27"]
    history = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        draw_iso = iso_date(row[0] if row else None)
        if not draw_iso:
            continue
        if draw_iso > lock_iso:
            continue
        codes = [number2(value) for value in row[1:28]]
        if len(codes) != 27:
            raise ValueError(f"{draw_iso}: canonical history row does not contain 27 codes")
        history.append((date.fromisoformat(draw_iso), codes))
    history.sort(key=lambda item: item[0])
    if not history or history[-1][0].isoformat() != lock_iso:
        raise ValueError(f"MB_History_27 does not end at data lock {lock_iso}")
    if len(history) < 181:
        raise ValueError("Not enough canonical history to build public methods")
    return history


def select_plan(wb) -> dict[str, Any]:
    rows = rows_as_dicts(wb["MAX2_Daily_Plan"])
    candidates = []
    for row in rows:
        if text(row.get("Method_ID")) != PRODUCTION_METHOD:
            continue
        target_iso = iso_date(row.get("Target_Date"))
        lock_iso = iso_date(row.get("Data_Lock"))
        if not target_iso or not lock_iso:
            continue
        if date.fromisoformat(lock_iso) != date.fromisoformat(target_iso) - timedelta(days=1):
            continue
        if not text(row.get("Source_Status")).startswith("PASS_27_LOCKED"):
            continue
        if text(row.get("Run_Status")) != "PUBLISHED_PROSPECTIVE":
            continue
        if boolish(row.get("Outcome_Known_At_Selection")):
            continue
        candidates.append((target_iso, row))
    if not candidates:
        raise ValueError("No valid prospective MAX2 Production row found")
    candidates.sort(key=lambda item: item[0])
    target_iso, row = candidates[-1]
    row = dict(row)
    row["_target_iso"] = target_iso
    row["_lock_iso"] = iso_date(row.get("Data_Lock"))
    return row


def settlement_rows(wb, lock_iso: str) -> list[dict[str, Any]]:
    rows = rows_as_dicts(wb["MAX2_Daily_Settlement"])
    result = []
    for row in rows:
        draw_iso = iso_date(row.get("Date"))
        if not draw_iso or draw_iso > lock_iso:
            continue
        if text(row.get("Result_Status")) != "SETTLED":
            continue
        if text(row.get("Execution_Status")) != "CONFIRMED_REAL":
            continue
        row = dict(row)
        row["_date_iso"] = draw_iso
        result.append(row)
    result.sort(key=lambda row: row["_date_iso"])
    return result


def backtest_summary(wb) -> dict[str, Any]:
    rows = rows_as_dicts(wb["MAX2_Backtest_2026_Monthly"])
    total = next((row for row in rows if text(row.get("Month")).startswith("TOTAL_")), None)
    august = next((row for row in rows if text(row.get("Month")).startswith("2026-08")), None)

    def compact(row):
        if not row:
            return None
        return {
            "period": text(row.get("Month")),
            "days": int(float(row.get("Days") or 0)),
            "hit_days": int(float(row.get("Hit_Days") or 0)),
            "hit_rate": float(row.get("Hit_Rate") or 0),
            "occurrences": int(float(row.get("Occurrences") or 0)),
            "status": text(row.get("Status")),
        }

    return {"total_2026": compact(total), "august_snapshot": compact(august)}


def history_map(history: list[tuple[date, list[str]]]) -> dict[str, list[str]]:
    return {draw.isoformat(): codes for draw, codes in history}


def record_for(draw_iso: str, outputs: list[str], actual: list[str]) -> dict[str, Any]:
    outputs = [number2(value) for value in outputs]
    counts = Counter(actual)
    hits = [{"number": code, "count": counts[code]} for code in outputs if counts[code] > 0]
    body = {"date": draw_iso, "recommended_numbers": outputs, "hits": hits, "status": "hit" if hits else "miss"}
    raw = json.dumps(body, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")
    body["record_hash"] = hashlib.sha256(raw).hexdigest()
    return body


def build_daily_records(history, settled, existing_proof, lock_iso):
    actual_by_date = history_map(history)
    by_date = {}
    month = (existing_proof.get("month_summary") or {}).get("daily_records") or []
    for item in month:
        draw_iso = iso_date(item.get("date"))
        if draw_iso and draw_iso < "2026-08-17":
            by_date[draw_iso] = item

    for draw_iso, outputs in LEGACY_BRIDGE.items():
        if draw_iso <= lock_iso and draw_iso in actual_by_date:
            by_date[draw_iso] = record_for(draw_iso, outputs, actual_by_date[draw_iso])

    for row in settled:
        draw_iso = row["_date_iso"]
        if draw_iso < "2026-08-19" or draw_iso not in actual_by_date:
            continue
        outputs = [number2(row.get("Code1")), number2(row.get("Code2"))]
        by_date[draw_iso] = record_for(draw_iso, outputs, actual_by_date[draw_iso])

    start_iso = f"{lock_iso[:7]}-01"
    current = date.fromisoformat(start_iso)
    end = date.fromisoformat(lock_iso)
    missing = []
    while current <= end:
        key = current.isoformat()
        if key not in by_date:
            missing.append(key)
        current += timedelta(days=1)
    if missing:
        raise ValueError(f"Official production history is incomplete through data lock: {missing}")
    return [by_date[key] for key in sorted(by_date) if start_iso <= key <= lock_iso]


def update_proofs(history, settled, public_payload, target_iso, lock_iso, now_iso):
    yesterday_path = ROOT / "ai-methods" / "yesterday-proof.json"
    historical_path = ROOT / "data" / "public-historical-proof.json"
    existing_yesterday = json.loads(yesterday_path.read_text(encoding="utf-8"))
    existing_historical = json.loads(historical_path.read_text(encoding="utf-8"))
    records = build_daily_records(history, settled, existing_yesterday, lock_iso)
    current_record = next(item for item in records if item["date"] == lock_iso)
    winning = [
        {"date": item["date"], "recommended_numbers": item["recommended_numbers"], "hits": item["hits"]}
        for item in records if item["status"] == "hit"
    ]
    hit_days = len(winning)
    total_days = len(records)
    existing_validation = existing_yesterday.get("historical_validation") or existing_historical.get("validation") or {}

    yesterday = {
        "schema_version": "MB_PUBLIC_YESTERDAY_PROOF_V3_PRODUCTION_AWARE",
        "date": lock_iso,
        "recommended_numbers": current_record["recommended_numbers"],
        "hits": current_record["hits"],
        "unique_hit_count": len(current_record["hits"]),
        "recommended_count": len(current_record["recommended_numbers"]),
        "total_occurrences": sum(int(item["count"]) for item in current_record["hits"]),
        "historical_validation": existing_validation,
        "month_summary": {
            "month": lock_iso[:7],
            "period_start": records[0]["date"],
            "period_end": lock_iso,
            "observed_days": total_days,
            "win_days": hit_days,
            "miss_days": total_days - hit_days,
            "daily_records": records,
            "winning_days": winning,
        },
        "production_transition": {
            "current_method": PRODUCTION_METHOD,
            "current_config": PRODUCTION_CONFIG,
            "max2_live_from": "2026-08-19",
            "policy": "Use the method that was officially locked before each draw; MAX2 is Production from 2026-08-19.",
        },
        "updated_at": now_iso,
    }
    yesterday_path.write_text(json.dumps(yesterday, ensure_ascii=False, separators=(",", ":")) + "\n", encoding="utf-8")

    historical_days = []
    for item in records:
        observed = []
        for hit in item["hits"]:
            count = int(hit["count"])
            observed.append(hit["number"] if count == 1 else f'{hit["number"]} × {count}')
        historical_days.append({
            "date": item["date"],
            "outputs": item["recommended_numbers"],
            "observed": observed,
            "status": item["status"],
        })
    validation = existing_historical.get("validation") or existing_validation
    historical = {
        "schema_version": "MB_PUBLIC_HISTORICAL_PROOF_V2_PRODUCTION_AWARE",
        "status": "COMPLETED_DATES_ONLY",
        "validation": validation,
        "recent_period": {
            "period_start": records[0]["date"],
            "period_end": lock_iso,
            "hit_days": hit_days,
            "total_days": total_days,
            "rate_pct": round(hit_days * 100 / total_days, 1) if total_days else 0,
            "days": historical_days,
        },
        "method_snapshot": {
            "target_date": target_iso,
            "data_lock": lock_iso,
            "source_status": "LOCKED_27_OF_27",
            "layers": [
                {"index": index + 1, "name": method["name"], "numbers": method["numbers"]}
                for index, method in enumerate(public_payload["methods"])
            ],
            "method_id": PRODUCTION_METHOD,
            "active_config": PRODUCTION_CONFIG,
            "paid_output_hidden": True,
        },
        "provenance": {
            "validation_source": "fixed historical validation snapshot",
            "method_snapshot_source": "current non-paid public methods + MB ALL Production status",
            "source": "XSMB_Source_2024_2026_MB_v1.3 / MB_History_27",
            "source_end": lock_iso,
            "policy": "HISTORICAL_COMPLETED_RECORDS_ONLY_NO_CURRENT_PAID_OUTPUTS",
            "generated_from": "Official per-day Production record; MAX2 from 2026-08-19",
            "updated_at": now_iso,
        },
    }
    historical_path.write_text(json.dumps(historical, ensure_ascii=False, separators=(",", ":")) + "\n", encoding="utf-8")


def write_status_files(wb, plan, history, settled, public_payload, now_iso):
    target_iso = plan["_target_iso"]
    lock_iso = plan["_lock_iso"]
    audit = json.loads((ROOT / "data" / "completed-draw-audit.json").read_text(encoding="utf-8"))
    audit_row = (audit.get("draws") or {}).get(lock_iso) or {}
    latest_codes = history[-1][1]
    latest_hash = hashlib.sha256("|".join(latest_codes).encode("utf-8")).hexdigest()
    audit_hash = text(audit_row.get("codes_sha256"))
    if audit_hash and audit_hash != latest_hash:
        raise ValueError(f"Completed draw audit hash mismatch for {lock_iso}")

    source = {
        "schema_version": "MB_SOURCE_ACCESS_V3_MAX2_SYNC",
        "status": "OK",
        "selected": "GOOGLE_SHEET_XLSX",
        "source_spreadsheet_id": SOURCE_SHEET_ID,
        "source_tab": "MB_History_27",
        "history_rows": len(history),
        "history_end": lock_iso,
        "latest_codes_sha256": latest_hash,
        "source_count": int(audit_row.get("source_count") or 1),
        "sources": audit_row.get("sources") or ["canonical_google_sheet"],
        "crosscheck_status": text(audit_row.get("status")) or "CANONICAL_LOCKED",
        "locked_at": now_iso,
        "target_date": target_iso,
        "production_method": PRODUCTION_METHOD,
        "milestone_requirement": "Ngày T chỉ dùng dữ liệu đến hết T-1.",
    }
    (ROOT / "data" / "source-access.json").write_text(json.dumps(source, ensure_ascii=False, separators=(",", ":")) + "\n", encoding="utf-8")

    ready = {
        "schema_version": "MB_PAID_REPORT_READINESS_V2_MAX2",
        "report_date": target_iso,
        "data_lock": lock_iso,
        "status": "PUBLISHED_PASS_PRIVATE",
        "outcome_known_at_selection": False,
        "updated_at": now_iso,
        "source": "MAX2_Daily_Plan -> private Paid_Report after daily review sync",
        "production_method": PRODUCTION_METHOD,
        "production_config": PRODUCTION_CONFIG,
        "note": "Public readiness only. Current Production codes are intentionally excluded.",
    }
    (ROOT / "data" / "paid-report-ready.json").write_text(json.dumps(ready, ensure_ascii=False, separators=(",", ":")) + "\n", encoding="utf-8")

    live = [row for row in settled if row["_date_iso"] >= "2026-08-19"]
    hit_days = sum(1 for row in live if int(float(row.get("Hit_Day") or 0)) > 0)
    cumulative = int(float(live[-1].get("Cumulative_PL_VND") or 0)) if live else 0
    site_status = {
        "schema_version": "MB_ALL_WEBSITE_DAILY_STATUS_V1",
        "project": "MB ALL",
        "production_method": PRODUCTION_METHOD,
        "production_config": PRODUCTION_CONFIG,
        "target_date": target_iso,
        "data_lock": lock_iso,
        "source_status": text(plan.get("Source_Status")),
        "run_status": text(plan.get("Run_Status")),
        "execution_status": text(plan.get("Execution_Status")),
        "outcome_known_at_selection": False,
        "public_ready": True,
        "paid_output_hidden": True,
        "updated_at": now_iso,
        "latest_completed_draw": lock_iso,
        "production_live": {
            "from_date": "2026-08-19",
            "settled_days": len(live),
            "hit_days": hit_days,
            "hit_rate_pct": round(hit_days * 100 / len(live), 1) if live else None,
            "cumulative_pl_vnd": cumulative,
        },
        "backtest": backtest_summary(wb),
        "invariants": {
            "data_lock_is_t_minus_1": True,
            "current_codes_not_public": True,
            "website_date_must_equal_target_date": True,
        },
    }
    (ROOT / "data" / "max2-site-status.json").write_text(json.dumps(site_status, ensure_ascii=False, separators=(",", ":")) + "\n", encoding="utf-8")


def sync(source_xlsx: Path) -> dict[str, Any]:
    wb = load_workbook(source_xlsx, read_only=True, data_only=True)
    try:
        plan = select_plan(wb)
        target_iso = plan["_target_iso"]
        lock_iso = plan["_lock_iso"]
        history = load_history(wb, lock_iso)
        public_payload = public_methods.build_payload(history, date.fromisoformat(target_iso))
        public_payload["production_context"] = {
            "project": "MB ALL",
            "production_method": PRODUCTION_METHOD,
            "paid_output_hidden": True,
        }
        public_payload["generated_at"] = datetime.now(VN).isoformat(timespec="seconds")
        (ROOT / "ai-methods" / "public-methods.json").write_text(json.dumps(public_payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        settled = settlement_rows(wb, lock_iso)
        now_iso = datetime.now(VN).isoformat(timespec="seconds")
        write_status_files(wb, plan, history, settled, public_payload, now_iso)
        update_proofs(history, settled, public_payload, target_iso, lock_iso, now_iso)
        return {
            "target_date": target_iso,
            "data_lock": lock_iso,
            "history_rows": len(history),
            "public_methods": len(public_payload["methods"]),
            "settled_production_days": len([row for row in settled if row["_date_iso"] >= "2026-08-19"]),
        }
    finally:
        wb.close()


def self_test() -> None:
    assert iso_date("21/08/2026") == "2026-08-21"
    assert iso_date("2026-08-21") == "2026-08-21"
    assert boolish("TRUE") and not boolish("FALSE")
    assert number2(0) == "00"
    assert number2(5.0) == "05"
    sample = record_for("2026-08-20", ["17", "50"], ["17", "17", "02"])
    assert sample["status"] == "hit"
    assert sample["hits"] == [{"number": "17", "count": 2}]
    print("MAX2_WEBSITE_SYNC_SELF_TEST_OK")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source-xlsx", type=Path)
    parser.add_argument("--self-test", action="store_true")
    args = parser.parse_args()
    if args.self_test:
        self_test()
        return
    if not args.source_xlsx or not args.source_xlsx.exists():
        raise SystemExit("--source-xlsx is required")
    result = sync(args.source_xlsx)
    print(json.dumps(result, ensure_ascii=False))


if __name__ == "__main__":
    main()
