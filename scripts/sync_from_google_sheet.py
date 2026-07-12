#!/usr/bin/env python3
"""Đồng bộ kết quả XSMB đã khóa từ Google Sheet lên website.

Nguyên tắc an toàn:
- Chỉ nhận kỳ có đủ 27 mã lô.
- Không cho dữ liệu nguồn lùi ngày; lệch nguồn cùng ngày đã cross-check thì dừng.
- Chỉ quyết toán lệnh thật ở trạng thái PENDING/SETTLED cùng ngày quay.
- Ghi sổ quyết toán theo ngày để chạy lại không cộng lãi/lỗ hai lần.
- Tự cập nhật số lệnh thắng, số lệnh thua và P/L lũy kế Xiên 2.
"""
from __future__ import annotations
import copy
import hashlib
import json
import os
import re
import tempfile
import urllib.request
from collections import Counter
from datetime import date, datetime, time, timedelta, timezone
from itertools import combinations
from pathlib import Path
from typing import Any
from openpyxl import load_workbook
ROOT = Path(__file__).resolve().parents[1]
DATA_FILE = ROOT / 'data' / 'current.json'
LEDGER_FILE = ROOT / 'data' / 'settlement-ledger.json'
SHEET_ID = os.getenv('GOOGLE_SHEET_ID', '1iVAfqmS-TvP02U8FtKSM2nr_7Dsd7qi2qEGnWV6IK7w')
EXPORT_URL = f'https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx'
SOURCE_XLSX_PATH = os.getenv('SOURCE_XLSX_PATH')
VN = timezone(timedelta(hours=7))
SOURCE_SHEETS = ('Raw_Results_2024_2026', 'Lo_Toan_Bang_2024_2026', 'Raw_Results_IMPORT', 'Raw_2Digits_IMPORT')

def now_vn() -> datetime:
    return datetime.now(VN)

def as_int(value: Any, default: int=0) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return default

def as_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        return (datetime(1899, 12, 30) + timedelta(days=float(value))).date()
    text = str(value or '').strip()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y'):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None

def code2(value: Any) -> str | None:
    if value is None or value == '':
        return None
    try:
        text = str(int(float(value)))
    except (TypeError, ValueError):
        text = str(value).strip()
    digits = ''.join((ch for ch in text if ch.isdigit()))
    return digits[-2:].zfill(2) if digits else None

def json_hash(value: Any) -> str:
    payload = json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(',', ':'))
    return hashlib.sha256(payload.encode('utf-8')).hexdigest()

def signed_vnd(value: int) -> str:
    return f'{value:+,}đ'.replace(',', '.')

def load_json(path: Path, default: dict[str, Any]) -> dict[str, Any]:
    if not path.exists():
        return copy.deepcopy(default)
    return json.loads(path.read_text(encoding='utf-8'))

def write_json(path: Path, value: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, ensure_ascii=False, separators=(',', ':')) + '\n', encoding='utf-8')

def latest_draw(xlsx_path: str) -> tuple[date, list[str], list[str]]:
    """Lấy kỳ mới nhất và bắt lỗi khi các tab mới nhất lệch nhau."""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    per_sheet: list[tuple[date, list[str], str]] = []
    try:
        for name in SOURCE_SHEETS:
            if name not in wb.sheetnames:
                continue
            ws = wb[name]
            best: tuple[date, list[str], str] | None = None
            for row in ws.iter_rows(values_only=True):
                if not row:
                    continue
                draw_date = as_date(row[0])
                codes = [code2(v) for v in row[1:28]]
                if draw_date and len(codes) == 27 and all(codes):
                    clean_codes = [str(x) for x in codes]
                    if best is None or draw_date > best[0]:
                        best = (draw_date, clean_codes, name)
            if best:
                per_sheet.append(best)
    finally:
        wb.close()
    if not per_sheet:
        raise RuntimeError('Không tìm thấy kỳ hợp lệ đủ 27 mã trong Google Sheet')
    newest = max((item[0] for item in per_sheet))
    latest = [item for item in per_sheet if item[0] == newest]
    variants = {tuple(item[1]) for item in latest}
    if len(variants) > 1:
        names = ', '.join((item[2] for item in latest))
        raise RuntimeError(f'Lệch dữ liệu giữa các tab cùng ngày {newest}: {names}')
    return (newest, latest[0][1], [item[2] for item in latest])

def normalise_pair(value: Any) -> str:
    parts = [p for p in re.split('[-–—/|,\\s]+', str(value or '').strip()) if p]
    if len(parts) != 2:
        raise ValueError(f'Cặp Xiên 2 không hợp lệ: {value!r}')
    left, right = (code2(parts[0]), code2(parts[1]))
    if not left or not right or left == right:
        raise ValueError(f'Cặp Xiên 2 không hợp lệ: {value!r}')
    return f'{left}-{right}'

def dedupe(values: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        if value not in seen:
            result.append(value)
            seen.add(value)
    return result

def find_group(doc: dict[str, Any], group_id: str) -> dict[str, Any] | None:
    for group in doc.get('groups') or []:
        if group.get('id') == group_id:
            return group
    return None

def extract_real_order(doc: dict[str, Any], draw_date: date) -> dict[str, Any] | None:
    order = doc.get('actual_order') or {}
    status = str(order.get('status', ''))
    eligible = 'REAL' in status and ('PENDING' in status or 'SETTLED' in status) or bool(order.get('pnl_included'))
    if order.get('date') == draw_date.isoformat() and eligible:
        return order
    group = find_group(doc, 'XIEN')
    group_status = str((group or {}).get('status', ''))
    target_date = as_date(doc.get('target_date'))
    if group and target_date == draw_date and ('REAL' in group_status) and ('PENDING' in group_status):
        synthetic = {'status': 'REAL_PENDING_FROM_XIEN_GROUP', 'date': draw_date.isoformat(), 'pnl_included': False, 'lo': {'numbers': []}, 'xien2': {'pairs': group.get('selected_numbers') or [], 'points_per_pair': group.get('points_per_pair', 100)}, 'note': 'Tự khôi phục lệnh thật từ khối XIEN cũ.'}
        doc['actual_order'] = synthetic
        return synthetic
    return None

def calculate_settlement(doc: dict[str, Any], order: dict[str, Any], draw_date: date, codes: list[str]) -> dict[str, Any]:
    """PER_CODE_MULTI_METHOD_SETTLEMENT_V2: settle each code at its confirmed stake."""
    freq = Counter(codes)
    stake = doc.get('stake_rule') or {}
    lo = order.get('lo') or {}
    numbers = dedupe([str(code2(x)).zfill(2) for x in lo.get('numbers', []) if code2(x)])
    fallback_points = as_int(lo.get('points_per_code'), as_int(stake.get('lo_points_per_code'), 50))
    raw_map = lo.get('points_by_code') or {}
    points_by_code = {
        number: as_int(raw_map.get(number), fallback_points)
        for number in numbers
    }
    if any(points <= 0 for points in points_by_code.values()):
        raise RuntimeError(f'Điểm theo mã không hợp lệ: {points_by_code}')
    lo_cost_pp = as_int(stake.get('lo_cost_per_point_vnd'), 23000)
    lo_pay_pp = as_int(stake.get('lo_payout_per_hit_point_vnd'), 80000)
    hits = {number: freq.get(number, 0) for number in numbers}
    lo_capital = sum(points_by_code[number] * lo_cost_pp for number in numbers)
    lo_payout = sum(hits[number] * points_by_code[number] * lo_pay_pp for number in numbers)
    lo_pnl = lo_payout - lo_capital
    uniform = len(set(points_by_code.values())) <= 1
    display_points = next(iter(points_by_code.values()), 0) if uniform else 0

    xien = order.get('xien2') or {}
    raw_pairs = xien.get('pairs') or []
    if not raw_pairs:
        base_numbers = [str(code2(x)).zfill(2) for x in xien.get('base_numbers', []) if code2(x)]
        if not base_numbers:
            base_numbers = numbers
        raw_pairs = [f'{a}-{b}' for a, b in combinations(dedupe(base_numbers), 2)] if xien.get('auto_combinations') else []
    pairs = dedupe([normalise_pair(pair) for pair in raw_pairs])
    xien_points = as_int(xien.get('points_per_pair'), as_int(stake.get('xien2_points_per_pair'), 100))
    pair_capital = as_int(stake.get('xien2_capital_per_pair_vnd'), 100000)
    pair_return = as_int(stake.get('xien2_gross_return_per_winning_pair_vnd'), 1600000)
    winning_pairs = [pair for pair in pairs if all(freq.get(leg, 0) > 0 for leg in pair.split('-', maxsplit=1))]
    xien_capital = len(pairs) * pair_capital
    xien_payout = len(winning_pairs) * pair_return
    xien_pnl = xien_payout - xien_capital
    return {
        'date': draw_date.isoformat(),
        'result_sha256': json_hash(codes),
        'result_codes': codes,
        'lo': {
            'numbers': numbers,
            'points_per_code': display_points,
            'points_mode': 'UNIFORM' if uniform else 'PER_CODE',
            'points_by_code': points_by_code,
            'hits': hits,
            'hits_total': sum(hits.values()),
            'capital_vnd': lo_capital,
            'payout_vnd': lo_payout,
            'pnl_vnd': lo_pnl,
        },
        'xien2': {
            'pairs': pairs,
            'points_per_pair': xien_points,
            'winning_pairs': winning_pairs,
            'wins': len(winning_pairs),
            'losses': len(pairs) - len(winning_pairs),
            'capital_per_pair_vnd': pair_capital,
            'gross_return_per_winning_pair_vnd': pair_return,
            'capital_vnd': xien_capital,
            'payout_vnd': xien_payout,
            'pnl_vnd': xien_pnl,
        },
        'total_capital_vnd': lo_capital + xien_capital,
        'total_payout_vnd': lo_payout + xien_payout,
        'total_pnl_vnd': lo_pnl + xien_pnl,
        'components': copy.deepcopy(order.get('components') or []),
    }

def default_ledger() -> dict[str, Any]:
    return {'schema_version': 'MB_SETTLEMENT_LEDGER_V1', 'baseline': {'xien2': {'orders': 0, 'wins': 0, 'losses': 0, 'pnl_vnd': 0}, 'source': 'EMPTY'}, 'settlements': {}}

def initialise_baseline(doc: dict[str, Any], ledger: dict[str, Any], record: dict[str, Any], already_included: bool) -> None:
    if ledger.get('baseline_initialized') or ledger.get('settlements'):
        return
    profile = (find_group(doc, 'XIEN') or {}).get('historical_profile') or {}
    orders = as_int(profile.get('orders'), 0)
    wins = as_int(profile.get('wins'), 0)
    losses = as_int(profile.get('losses'), max(orders - wins, 0))
    pnl = as_int(profile.get('pnl_vnd'), 0)
    if already_included:
        current = record['xien2']
        orders = max(orders - len(current['pairs']), 0)
        wins = max(wins - current['wins'], 0)
        losses = max(losses - current['losses'], 0)
        pnl -= current['pnl_vnd']
    ledger['baseline'] = {'xien2': {'orders': orders, 'wins': wins, 'losses': losses, 'pnl_vnd': pnl}, 'source': 'CURRENT_JSON_HISTORICAL_PROFILE'}
    ledger['baseline_initialized'] = True

def aggregate_xien(ledger: dict[str, Any]) -> dict[str, Any]:
    base = (ledger.get('baseline') or {}).get('xien2') or {}
    orders = as_int(base.get('orders'), 0)
    wins = as_int(base.get('wins'), 0)
    losses = as_int(base.get('losses'), max(orders - wins, 0))
    pnl = as_int(base.get('pnl_vnd'), 0)
    for record in (ledger.get('settlements') or {}).values():
        xien = record.get('xien2') or {}
        orders += len(xien.get('pairs') or [])
        wins += as_int(xien.get('wins'), 0)
        losses += as_int(xien.get('losses'), 0)
        pnl += as_int(xien.get('pnl_vnd'), 0)
    return {'orders': orders, 'wins': wins, 'losses': losses, 'win_rate': round(wins / orders, 6) if orders else 0.0, 'pnl_vnd': pnl, 'label': 'Sổ Xiên thực chiến'}

def prior_component_pnl(doc: dict[str, Any], order: dict[str, Any], previous: dict[str, Any] | None, component: str) -> int:
    """P/L đã thực sự được cộng vào current.json trước lần chạy này."""
    settlement = doc.get('settlement') or {}
    if order.get('pnl_included') and settlement.get('date') == order.get('date'):
        key = 'xien_pnl_vnd' if component == 'xien2' else 'lo_pnl_vnd'
        return as_int(settlement.get(key), 0)
    if order.get('pnl_included') and previous:
        return as_int((previous.get(component) or {}).get('pnl_vnd'), 0)
    return 0

def update_pnl_summary(doc: dict[str, Any], record: dict[str, Any], delta_lo: int, delta_xien: int) -> None:
    q = doc.setdefault('pnl_summary', {})
    total_delta = delta_lo + delta_xien
    for key in ('five_group_pnl_vnd', 'active_five_group_pnl_vnd'):
        if key in q:
            q[key] = as_int(q[key]) + delta_xien
    q['user_lo_pnl_vnd'] = as_int(q.get('user_lo_pnl_vnd'), 0) + delta_lo
    if 'active_all_real_pnl_vnd' in q:
        q['active_all_real_pnl_vnd'] = as_int(q['active_all_real_pnl_vnd']) + total_delta
    elif 'active_five_group_pnl_vnd' in q:
        q['active_all_real_pnl_vnd'] = as_int(q['active_five_group_pnl_vnd']) + as_int(q.get('user_lo_pnl_vnd'), 0)
    if 'grand_total_pnl_vnd' in q:
        q['grand_total_pnl_vnd'] = as_int(q['grand_total_pnl_vnd']) + total_delta
    elif 'active_all_real_pnl_vnd' in q:
        q['grand_total_pnl_vnd'] = as_int(q['active_all_real_pnl_vnd']) + as_int(q.get('archive_pnl_vnd'), 0)
    lo = record['lo']
    xien = record['xien2']
    stake_text = ', '.join(f"{code}×{lo['points_by_code'][code]}" for code in lo['numbers']) or 'Không có lô'
    if xien['pairs']:
        stake_text += f" + {len(xien['pairs'])} cặp Xiên ×{xien['points_per_pair']}"
    q.update({
        'confirmed_through': record['date'],
        'yesterday_date': record['date'],
        'yesterday_pnl_vnd': record['total_pnl_vnd'],
        'yesterday_hits': lo['hits_total'],
        'yesterday_stake': stake_text,
        'today_pending_capital_vnd': 0,
        'today_pending_order': f"Đã chốt P/L ngày: {signed_vnd(record['total_pnl_vnd'])}",
        'today_included': True,
    })

def update_ledger_snapshot(doc: dict[str, Any], aggregate: dict[str, Any]) -> None:
    snapshot = doc.setdefault('ledger_snapshot', {})
    q = doc.get('pnl_summary') or {}
    snapshot['confirmed_through'] = q.get('confirmed_through')
    snapshot['XIEN'] = aggregate['pnl_vnd']
    if 'active_five_group_pnl_vnd' in q:
        snapshot['FIVE_GROUPS'] = q['active_five_group_pnl_vnd']
    if 'user_lo_pnl_vnd' in q:
        snapshot['USER_LO'] = q['user_lo_pnl_vnd']
    if 'active_all_real_pnl_vnd' in q:
        snapshot['ALL_REAL'] = q['active_all_real_pnl_vnd']
    if 'archive_pnl_vnd' in q:
        snapshot['ARCHIVE'] = q['archive_pnl_vnd']
    if 'grand_total_pnl_vnd' in q:
        snapshot['GRAND_TOTAL'] = q['grand_total_pnl_vnd']

def update_xien_group(doc: dict[str, Any], record: dict[str, Any], aggregate: dict[str, Any]) -> None:
    group = find_group(doc, 'XIEN')
    if group is None:
        group = {'id': 'XIEN', 'label': 'Xiên 2', 'role': 'SỔ XIÊN RIÊNG'}
        doc.setdefault('groups', []).append(group)
    xien = record['xien2']
    pairs = xien['pairs']
    winning = set(xien['winning_pairs'])
    pair_capital = as_int(xien.get('capital_per_pair_vnd'), 100000)
    pair_return = as_int(xien.get('gross_return_per_winning_pair_vnd'), 1600000)
    group.update({'status': 'SETTLED_WIN' if xien['wins'] else 'SETTLED_LOSS', 'role': 'SỔ XIÊN RIÊNG', 'method': 'Tổ hợp Xiên 2 thực chiến', 'layer': f"{len(pairs)} cặp ×{xien['points_per_pair']} điểm", 'selected_numbers': pairs, 'selection': '|'.join(pairs), 'points': len(pairs) * xien['points_per_pair'], 'points_per_pair': xien['points_per_pair'], 'capital_vnd': xien['capital_vnd'], 'summary': f"{xien['wins']} thắng / {xien['losses']} thua; P/L ngày {signed_vnd(xien['pnl_vnd'])}.", 'reason': 'Cặp thắng: ' + ', '.join(xien['winning_pairs']) if xien['winning_pairs'] else 'Không có cặp nào đủ hai chân.', 'candidates': [{'code': pair, 'rank': index, 'gate': True, 'status': 'TRÚNG' if pair in winning else 'TRƯỢT', 'reason': f'P/L {signed_vnd(pair_return - pair_capital)}' if pair in winning else f'P/L {signed_vnd(-pair_capital)}'} for index, pair in enumerate(pairs, start=1)], 'historical_profile': aggregate})

def update_user_group(doc: dict[str, Any], record: dict[str, Any]) -> None:
    group = find_group(doc, 'USER_REAL')
    if group is None:
        return
    lo = record['lo']
    xien = record['xien2']
    group.update({'label': 'Kết quả thực chiến', 'status': 'SETTLED_WIN' if record['total_pnl_vnd'] > 0 else 'SETTLED_LOSS', 'layer': f"Quyết toán {datetime.strptime(record['date'], '%Y-%m-%d').strftime('%d/%m/%Y')}", 'summary': f"Lô {signed_vnd(lo['pnl_vnd'])}; Xiên {signed_vnd(xien['pnl_vnd'])}; tổng {signed_vnd(record['total_pnl_vnd'])}.", 'reason': f"Lô {lo['hits_total']} nháy; Xiên thắng {xien['wins']}/{len(xien['pairs'])} cặp."})

def full_settlement_display(doc: dict[str, Any], order: dict[str, Any], record: dict[str, Any]) -> None:
    lo = record['lo']
    xien = record['xien2']
    numbers = lo['numbers']
    pairs = xien['pairs']
    total_pnl = record['total_pnl_vnd']
    settled_at = now_vn().isoformat(timespec='seconds')
    order_lo = order.setdefault('lo', {})
    order_lo.update({
        'numbers': numbers,
        'points_per_code': lo['points_per_code'],
        'points_mode': lo['points_mode'],
        'points_by_code': lo['points_by_code'],
        'hits': lo['hits'],
        'code_count': len(numbers),
        'capital_vnd': lo['capital_vnd'],
        'payout_vnd': lo['payout_vnd'],
        'pnl_vnd': lo['pnl_vnd'],
    })
    order_xien = order.setdefault('xien2', {})
    order_xien.update({
        'pairs': pairs,
        'points_per_pair': xien['points_per_pair'],
        'winning_pairs': xien['winning_pairs'],
        'wins': xien['wins'],
        'losses': xien['losses'],
        'capital_vnd': xien['capital_vnd'],
        'payout_vnd': xien['payout_vnd'],
        'pnl_vnd': xien['pnl_vnd'],
    })
    order.update({
        'status': 'REAL_SETTLED_WIN' if total_pnl > 0 else 'REAL_SETTLED_LOSS' if total_pnl < 0 else 'REAL_SETTLED_EVEN',
        'settled_at': settled_at,
        'pnl_included': True,
        'total_capital_vnd': record['total_capital_vnd'],
        'total_payout_vnd': record['total_payout_vnd'],
        'total_pnl_vnd': total_pnl,
        'note': f"Lô {lo['hits_total']} nháy; Xiên thắng {xien['wins']}/{len(pairs)} cặp.",
    })
    doc['settlement'] = {
        'date': record['date'],
        'result_codes': record['result_codes'],
        'result_sha256': record['result_sha256'],
        'lo_hits_total': lo['hits_total'],
        'lo_points_by_code': lo['points_by_code'],
        'xien_wins': xien['wins'],
        'lo_pnl_vnd': lo['pnl_vnd'],
        'xien_pnl_vnd': xien['pnl_vnd'],
        'total_pnl_vnd': total_pnl,
    }
    doc.pop('pending_order', None)
    total_points = sum(lo['points_by_code'].values()) + xien['points_per_pair'] * len(pairs)
    doc['portfolio'] = {
        'decision': 'SETTLED_ACTUAL_ORDER',
        'tier': 'KẾT QUẢ ĐÃ KHÓA',
        'selection': '-'.join(numbers) if numbers else '|'.join(pairs),
        'title': f"KẾT QUẢ {datetime.strptime(record['date'], '%Y-%m-%d').strftime('%d/%m')}: {signed_vnd(total_pnl)}",
        'points': total_points,
        'capital_vnd': record['total_capital_vnd'],
        'payout_vnd': record['total_payout_vnd'],
        'pnl_vnd': total_pnl,
        'reason': order['note'],
        'pnl_status': 'SETTLED',
    }

    methods: list[dict[str, Any]] = []
    components = order.get('components') or []
    if components:
        pay_pp = as_int((doc.get('stake_rule') or {}).get('lo_payout_per_hit_point_vnd'), 80000)
        cost_pp = as_int((doc.get('stake_rule') or {}).get('lo_cost_per_point_vnd'), 23000)
        for component in components:
            component_numbers = [str(code2(x)).zfill(2) for x in component.get('selection', []) if code2(x)]
            component_map = {code: as_int((component.get('points_by_code') or {}).get(code), as_int(lo['points_by_code'].get(code))) for code in component_numbers}
            component_capital = sum(component_map[code] * cost_pp for code in component_numbers)
            component_payout = sum(lo['hits'].get(code, 0) * component_map[code] * pay_pp for code in component_numbers)
            component_pnl = component_payout - component_capital
            methods.append({
                'id': f"RESULT_{component.get('method_id', 'USER')}",
                'label': component.get('label') or component.get('method_id') or 'Lô thực chiến',
                'method': ' · '.join(f"{code}×{component_map[code]}" for code in component_numbers),
                'status': 'SETTLED_WIN' if component_pnl > 0 else 'SETTLED_LOSS' if component_pnl < 0 else 'SETTLED_EVEN',
                'points_per_code': next(iter(component_map.values()), 0) if len(set(component_map.values())) <= 1 else 'Theo mã',
                'code_count': len(component_numbers),
                'capital_vnd': component_capital,
                'payout_vnd': component_payout,
                'pnl_vnd': component_pnl,
                'numbers': [
                    {
                        'code': code,
                        'points': component_map[code],
                        'capital_vnd': component_map[code] * cost_pp,
                        'role': f"{lo['hits'].get(code, 0)} nháy",
                    }
                    for code in component_numbers
                ],
            })
    elif numbers:
        methods.append({
            'id': 'USER_LO',
            'label': 'Lô thực chiến',
            'method': ' · '.join(f"{code}×{lo['points_by_code'][code]}" for code in numbers),
            'status': order['status'],
            'points_per_code': lo['points_per_code'] or 'Theo mã',
            'code_count': len(numbers),
            'capital_vnd': lo['capital_vnd'],
            'numbers': [
                {
                    'code': number,
                    'points': lo['points_by_code'][number],
                    'capital_vnd': lo['points_by_code'][number] * as_int((doc.get('stake_rule') or {}).get('lo_cost_per_point_vnd'), 23000),
                    'role': f"{lo['hits'][number]} nháy",
                }
                for number in numbers
            ],
        })
    if pairs:
        methods.append({
            'id': 'USER_XIEN2',
            'label': 'Xiên 2 thực chiến',
            'method': f"{len(pairs)} cặp ×{xien['points_per_pair']} điểm",
            'status': 'SETTLED_WIN' if xien['wins'] else 'SETTLED_LOSS',
            'points_per_code': xien['points_per_pair'],
            'code_count': len(pairs),
            'capital_vnd': xien['capital_vnd'],
            'numbers': [
                {
                    'code': pair,
                    'points': xien['points_per_pair'],
                    'capital_vnd': as_int((doc.get('stake_rule') or {}).get('xien2_capital_per_pair_vnd'), 100000),
                    'role': 'Trúng' if pair in xien['winning_pairs'] else 'Trượt',
                }
                for pair in pairs
            ],
        })
    doc['top_signals'] = {
        'title': 'KẾT QUẢ THỰC CHIẾN ĐÃ CHỐT',
        'subtitle': f"Lô {signed_vnd(lo['pnl_vnd'])} · Xiên 2 {signed_vnd(xien['pnl_vnd'])} · Tổng {signed_vnd(total_pnl)}",
        'total_methods': len(methods),
        'total_numbers': f"{len(numbers)} số + {len(pairs)} cặp",
        'total_points': f"{sum(lo['points_by_code'].values())} lô + {xien['points_per_pair'] * len(pairs)} xiên",
        'total_capital_vnd': record['total_capital_vnd'],
        'note': order['note'],
        'methods': methods,
    }
    doc['top_signal_policy'] = {
        'mode': 'SETTLED_ACTUAL_ORDER',
        'system_signals_preserved_in_groups': True,
        'pnl_rule': 'Đã cộng kết quả đúng một lần vào sổ xác nhận.',
        'settlement_rule_version': 'PER_CODE_MULTI_METHOD_SETTLEMENT_V2',
    }

def settle_order(doc: dict[str, Any], ledger: dict[str, Any], draw_date: date, codes: list[str]) -> bool:
    order = extract_real_order(doc, draw_date)
    if not order:
        return False
    record = calculate_settlement(doc, order, draw_date, codes)
    settlements = ledger.setdefault('settlements', {})
    previous = settlements.get(draw_date.isoformat())
    current_settlement = doc.get('settlement') or {}
    already_included = bool(order.get('pnl_included')) and current_settlement.get('date') == draw_date.isoformat()
    initialise_baseline(doc, ledger, record, already_included and previous is None)
    old_lo = prior_component_pnl(doc, order, previous, 'lo')
    old_xien = prior_component_pnl(doc, order, previous, 'xien2')
    delta_lo = record['lo']['pnl_vnd'] - old_lo
    delta_xien = record['xien2']['pnl_vnd'] - old_xien
    result_changed = not already_included or as_int(current_settlement.get('lo_pnl_vnd')) != record['lo']['pnl_vnd'] or as_int(current_settlement.get('xien_pnl_vnd')) != record['xien2']['pnl_vnd'] or (current_settlement.get('result_sha256') not in (None, record['result_sha256'])) or (current_settlement.get('result_codes') != record['result_codes'])
    settlements[draw_date.isoformat()] = record
    aggregate = aggregate_xien(ledger)
    update_xien_group(doc, record, aggregate)
    update_pnl_summary(doc, record, delta_lo, delta_xien)
    update_ledger_snapshot(doc, aggregate)
    if result_changed:
        full_settlement_display(doc, order, record)
        update_user_group(doc, record)
    else:
        order['pnl_included'] = True
    return True

def build_data_snapshot(draw_date: date, codes: list[str], source_tabs: list[str]) -> dict[str, Any]:
    counts = Counter(codes)
    repeats = [f'{code}×{count}' for code, count in sorted(counts.items()) if count > 1]
    code_hash = json_hash(codes)
    repeat_code_count = sum((1 for count in counts.values() if count >= 2))
    max_frequency = max(counts.values())
    return {'locked_through': draw_date.isoformat(), 'status': 'LOCKED_AUTO_GOOGLE_SHEET', 'count': 27, 'unique': len(counts), 'repeat2_count': repeat_code_count, 'repeat_codes': repeats, 'max_frequency': max_frequency, 'noise_status': 'STRONG_NOISE' if max_frequency >= 3 or repeat_code_count >= 3 else 'NORMAL', 'latest_27_codes': codes, 'source_tabs': source_tabs, 'source_label': f"Google Sheet XSMB · khóa tự động {draw_date.strftime('%d/%m/%Y')} · đủ 27/27", 'sha256': code_hash}

def sync_data(doc: dict[str, Any], draw_date: date, codes: list[str], source_tabs: list[str]) -> str:
    existing = doc.get('data') or {}
    locked = as_date(existing.get('locked_through'))
    existing_codes = existing.get('latest_27_codes') or []
    if locked and draw_date < locked:
        return 'SOURCE_OLDER_SKIP'
    if locked == draw_date and existing_codes and (existing_codes != codes):
        status = str(existing.get('status', ''))
        if 'CROSSCHECKED' in status:
            raise RuntimeError(f'Google Sheet lệch bộ 27 mã đã cross-check ngày {draw_date}; dừng tự động để kiểm tra')
    if locked != draw_date or existing_codes != codes or len(existing_codes) != 27:
        doc['data'] = build_data_snapshot(draw_date, codes, source_tabs)
        doc['target_date'] = draw_date.isoformat()
        doc['data_snapshot_id'] = f"AUTO_{draw_date.strftime('%Y%m%d')}_{json_hash(codes)[:10].upper()}"
        doc['valid_until'] = datetime.combine(draw_date + timedelta(days=1), time(13, 59, 59), tzinfo=VN).isoformat()
        return 'UPDATED'
    return 'UNCHANGED'

def obtain_draw() -> tuple[date, list[str], list[str]]:
    if SOURCE_XLSX_PATH:
        return latest_draw(SOURCE_XLSX_PATH)
    with tempfile.NamedTemporaryFile(suffix='.xlsx') as tmp:
        req = urllib.request.Request(EXPORT_URL, headers={'User-Agent': 'MB-Daily-Control/2.0'})
        with urllib.request.urlopen(req, timeout=90) as response:
            tmp.write(response.read())
            tmp.flush()
        return latest_draw(tmp.name)

def main() -> None:
    if not DATA_FILE.exists():
        raise RuntimeError(f'Thiếu tệp dữ liệu website: {DATA_FILE}')
    doc = load_json(DATA_FILE, {})
    ledger = load_json(LEDGER_FILE, default_ledger())
    before_doc = copy.deepcopy(doc)
    before_ledger = copy.deepcopy(ledger)
    draw_date, codes, source_tabs = obtain_draw()
    if draw_date > now_vn().date():
        raise RuntimeError(f'Nguồn có ngày tương lai {draw_date}; dừng tự động')
    data_status = sync_data(doc, draw_date, codes, source_tabs)
    if data_status == 'SOURCE_OLDER_SKIP':
        print(f'Nguồn mới nhất {draw_date} cũ hơn dữ liệu website; giữ nguyên, không ghi đè.')
        return
    settled = settle_order(doc, ledger, draw_date, codes)
    changed_doc = doc != before_doc
    changed_ledger = ledger != before_ledger
    if not changed_doc and (not changed_ledger):
        print(f'Không có thay đổi: kỳ {draw_date}, settlement={settled}.')
        return
    timestamp = now_vn().isoformat(timespec='seconds')
    if changed_doc:
        doc['generated_at'] = timestamp
        doc['automation'] = {'status': 'OK', 'source': 'GOOGLE_SHEET_XLSX', 'last_updated_at': timestamp, 'locked_through': draw_date.isoformat(), 'settlement_checked': settled}
        write_json(DATA_FILE, doc)
    if changed_ledger:
        ledger['updated_at'] = timestamp
        write_json(LEDGER_FILE, ledger)
    print(f'Đã đồng bộ {draw_date}: data={data_status}, settlement={settled}, current_json={changed_doc}, ledger={changed_ledger}.')
if __name__ == '__main__':
    main()
