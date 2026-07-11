#!/usr/bin/env python3
"""Build a canonical XSMB history workbook without depending on Google export.

The repository carries a compressed, cross-checked 2024+ history bootstrap. After
18:30 Vietnam time, the script fetches the current draw from several independent
public result pages, requires at least two exact 27-code matches, appends the draw,
and emits an XLSX consumed by the settlement and next-day planning engines.
"""
from __future__ import annotations

import argparse
import base64
import bz2
import hashlib
import html as html_lib
import json
import os
import re
import sys
import urllib.error
import urllib.request
from collections import Counter, defaultdict
from datetime import date, datetime, time, timedelta, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
BOOTSTRAP_DIR = DATA_DIR / "history-bootstrap"
HISTORY_FILE = DATA_DIR / "history-27.bz2.b64"
ACCESS_FILE = DATA_DIR / "source-access.json"
X2_STATE_FILE = DATA_DIR / "x2-pilot-state.json"
VN = timezone(timedelta(hours=7))
EXPECTED_BOOTSTRAP_RAW_SHA256 = "066e6e1020eaa5608c62c8db40e4583ea5526e31366471fd88305cd769a59ed8"
MIN_CROSSCHECK_SOURCES = 2

PRIZES = [
    ("DB", 5, 1),
    ("G1", 5, 1),
    ("G2", 5, 2),
    ("G3", 5, 6),
    ("G4", 4, 4),
    ("G5", 4, 6),
    ("G6", 3, 3),
    ("G7", 2, 4),
]

LABELS = {
    "DB": re.compile(
        r"(?:G\s*\.\s*(?:ĐB|DB)|Giải\s*(?:ĐB|DB|Đặc\s*biệt)|Đặc\s*biệt|(?<![\w.])ĐB(?!\w))",
        re.IGNORECASE,
    ),
    "G1": re.compile(r"(?:G\s*\.\s*1|Giải\s*(?:nhất|1)(?!\d))", re.IGNORECASE),
    "G2": re.compile(r"(?:G\s*\.\s*2|Giải\s*(?:nhì|hai|2)(?!\d))", re.IGNORECASE),
    "G3": re.compile(r"(?:G\s*\.\s*3|Giải\s*(?:ba|3)(?!\d))", re.IGNORECASE),
    "G4": re.compile(r"(?:G\s*\.\s*4|Giải\s*(?:tư|bốn|4)(?!\d))", re.IGNORECASE),
    "G5": re.compile(r"(?:G\s*\.\s*5|Giải\s*(?:năm|5)(?!\d))", re.IGNORECASE),
    "G6": re.compile(r"(?:G\s*\.\s*6|Giải\s*(?:sáu|6)(?!\d))", re.IGNORECASE),
    "G7": re.compile(r"(?:G\s*\.\s*7|Giải\s*(?:bảy|7)(?!\d))", re.IGNORECASE),
}


def now_vn() -> datetime:
    return datetime.now(VN)


def canonical_json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def write_json_if_changed(path: Path, value: Any) -> bool:
    text = json.dumps(value, ensure_ascii=False, separators=(",", ":")) + "\n"
    if path.exists() and path.read_text(encoding="utf-8") == text:
        return False
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")
    return True


def bootstrap_b64() -> str:
    parts = sorted(BOOTSTRAP_DIR.glob("history-27-2024.bz2.b64.part-*"))
    if not parts:
        raise RuntimeError("Thiếu các phần history bootstrap trong data/history-bootstrap")
    return "".join(part.read_text(encoding="utf-8").strip() for part in parts)


def load_history() -> dict[str, Any]:
    if HISTORY_FILE.exists():
        encoded = HISTORY_FILE.read_text(encoding="utf-8").strip()
        source = HISTORY_FILE.name
    else:
        encoded = bootstrap_b64()
        source = "history-bootstrap"
    try:
        raw = bz2.decompress(base64.b64decode(encoded, validate=True))
    except Exception as exc:
        raise RuntimeError(f"History bootstrap hỏng: {exc!r}") from exc
    digest = hashlib.sha256(raw).hexdigest()
    doc = json.loads(raw.decode("utf-8"))
    rows = doc.get("rows") or []
    if source == "history-bootstrap" and digest != EXPECTED_BOOTSTRAP_RAW_SHA256:
        raise RuntimeError(f"Sai SHA256 history bootstrap: {digest}")
    validate_rows(rows)
    return doc


def save_history(doc: dict[str, Any]) -> bool:
    raw = json.dumps(doc, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
    encoded = base64.b64encode(bz2.compress(raw, compresslevel=9)).decode("ascii") + "\n"
    if HISTORY_FILE.exists() and HISTORY_FILE.read_text(encoding="utf-8") == encoded:
        return False
    HISTORY_FILE.write_text(encoded, encoding="utf-8")
    return True


def validate_rows(rows: list[Any]) -> None:
    if len(rows) < 500:
        raise RuntimeError(f"History chỉ có {len(rows)} kỳ; không đủ")
    seen: dict[str, list[str]] = {}
    previous: date | None = None
    for idx, row in enumerate(rows):
        if not isinstance(row, list) or len(row) != 28:
            raise RuntimeError(f"History row {idx} không đủ date + 27 mã")
        try:
            day = date.fromisoformat(str(row[0]))
        except Exception as exc:
            raise RuntimeError(f"History row {idx} sai ngày: {row[0]!r}") from exc
        codes = [str(value).zfill(2) for value in row[1:]]
        if any(not re.fullmatch(r"\d{2}", code) for code in codes):
            raise RuntimeError(f"History row {idx} có mã sai")
        if previous is not None and day <= previous:
            raise RuntimeError(f"History không tăng dần tại {day}")
        old = seen.get(day.isoformat())
        if old is not None and old != codes:
            raise RuntimeError(f"History lệch hai bộ mã ngày {day}")
        seen[day.isoformat()] = codes
        previous = day


def strip_html(raw_html: str) -> str:
    text = re.sub(r"(?is)<(?:script|style|noscript)[^>]*>.*?</(?:script|style|noscript)>", " ", raw_html)
    text = re.sub(r"(?s)<!--.*?-->", " ", text)
    text = re.sub(r"(?s)<[^>]+>", " ", text)
    text = html_lib.unescape(text)
    text = text.replace("\xa0", " ")
    return re.sub(r"\s+", " ", text).strip()


def parse_prizes(raw_html: str) -> list[str]:
    plain = strip_html(raw_html)
    db_candidates = list(LABELS["DB"].finditer(plain))
    for db_match in db_candidates:
        cursor = db_match.end()
        blocks: list[tuple[str, str, int, int]] = []
        ok = True
        current_match = db_match
        for prize_index, (key, width, count) in enumerate(PRIZES):
            if prize_index == 0:
                current_match = db_match
            else:
                candidate = LABELS[key].search(plain, cursor, min(len(plain), cursor + 800))
                if candidate is None:
                    ok = False
                    break
                current_match = candidate
            next_start = min(len(plain), current_match.end() + 800)
            if prize_index + 1 < len(PRIZES):
                next_key = PRIZES[prize_index + 1][0]
                nxt = LABELS[next_key].search(plain, current_match.end(), next_start)
                if nxt is None:
                    ok = False
                    break
                block = plain[current_match.end() : nxt.start()]
                cursor = nxt.start()
            else:
                block = plain[current_match.end() : current_match.end() + 160]
            numbers = re.findall(rf"(?<!\d)\d{{{width}}}(?!\d)", block)
            if len(numbers) != count:
                ok = False
                break
            blocks.append((key, block, width, count))
        if not ok:
            continue
        prizes: list[str] = []
        for key, block, width, count in blocks:
            numbers = re.findall(rf"(?<!\d)\d{{{width}}}(?!\d)", block)[:count]
            prizes.extend(numbers)
        if len(prizes) == 27:
            return [number[-2:] for number in prizes]
    raise ValueError("Không tách được đủ 27 giải theo cấu trúc DB→G7")


def result_urls(draw_date: date) -> list[tuple[str, str]]:
    dmy = draw_date.strftime("%d-%m-%Y")
    dmy_slash = draw_date.strftime("%d/%m/%Y")
    return [
        ("xosodaiphat", f"https://xosodaiphat.com/xsmb-{dmy}.html"),
        ("xosothienphu", f"https://xosothienphu.vn/xsmb-{dmy}.html"),
        ("xoso.com.vn", f"https://xoso.com.vn/xsmb-{dmy}.html"),
        ("minhngoc", f"https://www.minhngoc.net.vn/ket-qua-xo-so/mien-bac/{dmy}.html"),
        ("kqxs", f"https://kqxs.vn/mien-bac/xsmb-{dmy}"),
        ("ketqua", f"https://ketqua.net/xo-so-truyen-thong.php?ngay={dmy}"),
    ]


def fetch_page(url: str) -> str:
    request = urllib.request.Request(
        url,
        headers={
            "User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 Chrome/126 Safari/537.36",
            "Accept": "text/html,application/xhtml+xml",
            "Accept-Language": "vi-VN,vi;q=0.9,en;q=0.7",
        },
    )
    with urllib.request.urlopen(request, timeout=30) as response:
        body = response.read(2_000_000)
    return body.decode("utf-8", "replace")


def crosscheck_draw(draw_date: date) -> tuple[list[str], list[dict[str, Any]], list[dict[str, Any]]]:
    groups: dict[tuple[str, ...], list[dict[str, Any]]] = defaultdict(list)
    failures: list[dict[str, Any]] = []
    for source, url in result_urls(draw_date):
        try:
            raw = fetch_page(url)
            codes = parse_prizes(raw)
            record = {
                "source": source,
                "url": url,
                "codes_sha256": hashlib.sha256("|".join(codes).encode()).hexdigest(),
            }
            groups[tuple(codes)].append(record)
        except Exception as exc:
            failures.append({"source": source, "url": url, "error": repr(exc)[:300]})
    if not groups:
        raise RuntimeError(f"Không nguồn nào tách được kết quả {draw_date}: {failures}")
    ranked = sorted(groups.items(), key=lambda item: (len(item[1]), item[0]), reverse=True)
    winning_codes, sources = ranked[0]
    if len(sources) < MIN_CROSSCHECK_SOURCES:
        variants = {hashlib.sha256("|".join(key).encode()).hexdigest()[:12]: [r["source"] for r in value] for key, value in groups.items()}
        raise RuntimeError(f"Kết quả {draw_date} chưa đủ {MIN_CROSSCHECK_SOURCES} nguồn khớp: {variants}; failures={failures}")
    return list(winning_codes), sources, failures


def maybe_append_public_draws(doc: dict[str, Any], force_date: date | None = None) -> tuple[bool, list[dict[str, Any]]]:
    rows: list[list[str]] = doc["rows"]
    latest = date.fromisoformat(rows[-1][0])
    now = now_vn()
    cutoff_reached = now.time() >= time(18, 30)
    target = force_date or (now.date() if cutoff_reached else latest)
    if target <= latest:
        return False, []
    appended = False
    audits: list[dict[str, Any]] = []
    day = latest + timedelta(days=1)
    while day <= target:
        try:
            codes, sources, failures = crosscheck_draw(day)
        except Exception as exc:
            if day == target and (force_date is not None or cutoff_reached):
                raise
            audits.append({"date": day.isoformat(), "status": "NO_DRAW_OR_NOT_LOCKED", "error": repr(exc)[:500]})
            day += timedelta(days=1)
            continue
        old = next((row[1:] for row in rows if row[0] == day.isoformat()), None)
        if old is not None and old != codes:
            raise RuntimeError(f"Public crosscheck lệch history đã khóa ngày {day}")
        if old is None:
            rows.append([day.isoformat(), *codes])
            rows.sort(key=lambda row: row[0])
            appended = True
        audits.append({
            "date": day.isoformat(),
            "status": "LOCKED_CROSSCHECKED_PUBLIC",
            "sources": sources,
            "failures": failures,
            "codes_sha256": hashlib.sha256("|".join(codes).encode()).hexdigest(),
        })
        day += timedelta(days=1)
    validate_rows(rows)
    return appended, audits


def build_workbook(doc: dict[str, Any], output: Path) -> None:
    rows: list[list[str]] = doc["rows"]
    wb = Workbook()
    ws = wb.active
    ws.title = "MB_History_27"
    header = ["date", *[f"L{i:02d}" for i in range(1, 28)]]
    ws.append(header)
    for row in rows:
        ws.append(row)
    for title in ("Raw_Results_IMPORT", "Raw_2Digits_IMPORT"):
        copy_ws = wb.create_sheet(title)
        copy_ws.append(header)
        for row in rows:
            copy_ws.append(row)
    x2 = wb.create_sheet("X2_Live_Log_v1")
    x2_header = [
        "Date", "Data_Status", "Prior_Controller", "Prev_Repeat2", "Prev_MaxFreq",
        "X2_Tier", "Pair", "Main", "Cover", "Gan_Main", "Gan_Cover", "H5_Pair",
        "H21_Main", "H21_Cover", "H60_Main", "H60_Cover", "H90_Main", "H90_Cover",
        "Decision", "Real_Confirmed", "Points_Per_Code", "Capital_VND", "Main_Hits",
        "Cover_Hits", "Total_Hits", "P/L_Day_VND", "Cumulative_VND", "Peak_VND",
        "Drawdown_VND", "Loss_Streak", "Brake_Remaining", "Pilot_Status", "Notes",
    ]
    x2.append(x2_header)
    if X2_STATE_FILE.exists():
        state = json.loads(X2_STATE_FILE.read_text(encoding="utf-8"))
        x2.append([
            state.get("date", rows[-1][0]), "REPO_STATE", "", 0, 0, "", "", "", "", 0, 0,
            0, 0, 0, 0, 0, 0, 0, "STATE", False, 15, 0, 0, 0, 0, 0,
            state.get("cumulative_vnd", 0), state.get("peak_vnd", 0),
            state.get("drawdown_vnd", 0), state.get("loss_streak", 0),
            state.get("brake_remaining", 0), state.get("status", "ACTIVE_GUARDED"),
            "Persistent repository X2 brake state",
        ])
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def self_test() -> None:
    doc = load_history()
    assert doc["rows"][0][0] == "2024-01-01"
    assert doc["rows"][-1][0] == "2026-07-11"
    sample = """
      <div>G.ĐB09401</div><div>G.136061</div><div>G.277252 60057</div>
      <div>G.351690 28065 93903 75131 65832 12023</div>
      <div>G.43626 1683 2414 9774</div><div>G.59198 1500 3618 8389 9640 0250</div>
      <div>G.6425 731 475</div><div>G.706 26 73 72</div>
    """
    expected = ["01", "61", "52", "57", "90", "65", "03", "31", "32", "23", "26", "83", "14", "74", "98", "00", "18", "89", "40", "50", "25", "31", "75", "06", "26", "73", "72"]
    assert parse_prizes(sample) == expected
    with Path("/tmp/public-history-self-test.xlsx").open("wb"):
        pass
    build_workbook(doc, Path("/tmp/public-history-self-test.xlsx"))
    print("PUBLIC_SOURCE_SELF_TEST_OK", len(doc["rows"]), doc["rows"][-1][0])


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", default="/tmp/xsmb-source.xlsx")
    parser.add_argument("--force-date")
    parser.add_argument("--self-test", action="store_true")
    args = parser.parse_args()
    if args.self_test:
        self_test()
        return
    force_date = date.fromisoformat(args.force_date) if args.force_date else None
    doc = load_history()
    before_latest = doc["rows"][-1][0]
    changed, audits = maybe_append_public_draws(doc, force_date)
    history_changed = save_history(doc)
    build_workbook(doc, Path(args.output))
    latest = doc["rows"][-1]
    access = {
        "schema_version": "MB_SOURCE_ACCESS_V2",
        "status": "LOCKED_CROSSCHECKED_PUBLIC" if audits else "BOOTSTRAP_HISTORY_READY",
        "selected": "PUBLIC_MULTI_SOURCE_CROSSCHECK" if audits else "REPOSITORY_HISTORY_BOOTSTRAP",
        "history_start": doc["rows"][0][0],
        "history_end": latest[0],
        "history_rows": len(doc["rows"]),
        "latest_codes_sha256": hashlib.sha256("|".join(latest[1:]).encode()).hexdigest(),
        "appended": changed,
        "audits": audits,
        "milestone_requirement": "A1/X2/X3 candidates must include earliest date, condition, and type",
    }
    write_json_if_changed(ACCESS_FILE, access)
    print(canonical_json({"before_latest": before_latest, "after_latest": latest[0], "history_changed": history_changed, "output": args.output, "audits": len(audits)}))


if __name__ == "__main__":
    main()
