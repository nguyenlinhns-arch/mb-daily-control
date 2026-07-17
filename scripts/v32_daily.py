#!/usr/bin/env python3
"""Prepare and finalize the fail-closed 06:00 V32 daily transaction."""
from __future__ import annotations

import argparse
from collections import Counter
from datetime import date, datetime, timedelta, timezone
from hashlib import sha256
from html import escape
import json
import math
import os
from pathlib import Path
import re
import shutil
import subprocess
import sys
from typing import Any

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
ENGINE = ROOT / "engine_v32"
TEMPLATE = ROOT / "templates" / "v32-dashboard.html"
DATA = ROOT / "data"
STATE_FILE = DATA / "v32-state.json"
PLAN_DIR = DATA / "v32-plans"
SETTLEMENT_DIR = DATA / "v32-settlements"
TX_DIR = DATA / "v32-transactions"
MANIFEST = ENGINE / "manifest.json"

VN = timezone(timedelta(hours=7))
COST_PER_POINT = 23_000
PAYOUT_PER_HIT_POINT = 80_000
METHOD = "R868Y26 + Gate315 + Empirical Overdue Swap075"
METHOD_WITH_STAKE = METHOD + " / A6_B8"
PIPELINE_VERSION = "MB_V32_DAILY_0600_TXN_V2"
PEOPLE = ("p1", "p2", "p3", "p4", "p5")
WEEKDAYS = (
    "Thứ Hai", "Thứ Ba", "Thứ Tư", "Thứ Năm",
    "Thứ Sáu", "Thứ Bảy", "Chủ Nhật",
)


class PipelineError(RuntimeError):
    pass


def canonical(value: Any) -> bytes:
    return json.dumps(
        value, ensure_ascii=False, sort_keys=True, separators=(",", ":")
    ).encode("utf-8")


def digest(value: Any) -> str:
    return sha256(canonical(value)).hexdigest()


def read_json(path: Path) -> dict:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (FileNotFoundError, json.JSONDecodeError) as exc:
        raise PipelineError(f"Không đọc được JSON bắt buộc: {path}") from exc


def write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(value, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def write_text(path: Path, value: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(value, encoding="utf-8")


def as_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        return (datetime(1899, 12, 30) + timedelta(days=float(value))).date()
    text = str(value or "").strip()[:10]
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def code2(value: Any) -> str:
    if isinstance(value, bool):
        raise PipelineError(f"Mã không hợp lệ: {value}")
    if isinstance(value, float):
        if not math.isfinite(value) or not value.is_integer():
            raise PipelineError(f"Mã không nguyên: {value}")
        number = int(value)
    elif isinstance(value, int):
        number = value
    else:
        text = str(value).strip()
        if not re.fullmatch(r"\d{1,2}", text):
            raise PipelineError(f"Mã không hợp lệ: {value}")
        number = int(text)
    if not 0 <= number <= 99:
        raise PipelineError(f"Mã ngoài khoảng 00-99: {value}")
    return f"{number:02d}"


def load_history(path: Path) -> dict[date, list[str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if "MB_History_27" not in wb.sheetnames:
            raise PipelineError("Nguồn thiếu tab MB_History_27")
        rows: dict[date, list[str]] = {}
        for row in wb["MB_History_27"].iter_rows(values_only=True):
            if not row:
                continue
            day = as_date(row[0])
            if day is None:
                continue
            values = []
            for raw in row[1:28]:
                if raw is None or raw == "":
                    continue
                if isinstance(raw, (int, float)):
                    if (
                        isinstance(raw, bool) or not math.isfinite(float(raw))
                        or float(raw) < 0 or not float(raw).is_integer()
                    ):
                        raise PipelineError(f"Kết quả nguồn không nguyên/dương: {raw}")
                    values.append(f"{int(raw) % 100:02d}")
                    continue
                text = str(raw).strip()
                if not re.fullmatch(r"\d+", text):
                    raise PipelineError(f"Kết quả nguồn sai định dạng: {raw}")
                values.append(f"{int(text) % 100:02d}")
            if len(values) == 27:
                if day in rows and rows[day] != values:
                    raise PipelineError(f"Nguồn xung đột tại {day}")
                rows[day] = values
    finally:
        wb.close()
    if not rows:
        raise PipelineError("Nguồn không có kỳ đủ 27/27")
    return rows


def write_truncated_engine_workbook(
    history: dict[date, list[str]], source_end: date, destination: Path
) -> None:
    if source_end not in history:
        raise PipelineError(f"Chưa có kết quả đủ 27/27 ngày {source_end}")
    wb = Workbook()
    ws = wb.active
    ws.title = "MB_History_27"
    ws.append(["Date"] + [f"L{i:02d}" for i in range(1, 28)])
    for day in sorted(history):
        if day <= source_end:
            ws.append([day.isoformat()] + history[day])
    destination.parent.mkdir(parents=True, exist_ok=True)
    wb.save(destination)


def load_plan(day: date) -> dict:
    path = PLAN_DIR / f"{day.isoformat()}.json"
    if not path.exists():
        raise PipelineError(f"Thiếu lệnh đã khóa ngày {day}")
    plan = read_json(path)
    if plan.get("target_date") != day.isoformat():
        raise PipelineError("Ngày trong lệnh đã khóa không khớp tên file")
    codes = [code2(value) for value in plan.get("codes", [])]
    points = {code2(k): int(v) for k, v in plan.get("points_by_code", {}).items()}
    if not codes or len(codes) != len(set(codes)) or set(codes) != set(points):
        raise PipelineError("Lệnh trước không hợp lệ hoặc có mã trùng")
    width = len(codes)
    if not 1 <= width <= 12:
        raise PipelineError("Lệnh trước vi phạm độ rộng 1-12")
    expected = 30 if width <= 6 else 25 if width <= 8 else 20
    if any(value != expected for value in points.values()):
        raise PipelineError("Lệnh trước có điểm âm/hỗn hợp hoặc sai A6_B8")
    if plan.get("points_per_code") != expected:
        raise PipelineError("Lệnh trước sai bậc điểm A6_B8")
    total_points = sum(points.values())
    if plan.get("total_points") != total_points:
        raise PipelineError("Lệnh trước sai tổng điểm")
    if plan.get("total_capital_vnd") != total_points * COST_PER_POINT:
        raise PipelineError("Lệnh trước sai vốn")
    plan["codes"] = codes
    plan["points_by_code"] = points
    return plan


def settle(plan: dict, draw: list[str]) -> dict:
    counts = Counter(draw)
    hits = {code: counts.get(code, 0) for code in plan["codes"]}
    hit_units = sum(hits.values())
    total_points = sum(plan["points_by_code"].values())
    capital = total_points * COST_PER_POINT
    payout = sum(
        hits[code] * plan["points_by_code"][code] * PAYOUT_PER_HIT_POINT
        for code in plan["codes"]
    )
    pnl = payout - capital
    return {
        "date": plan["target_date"],
        "status": "SETTLED_VERIFIED_27_OF_27",
        "method": METHOD_WITH_STAKE,
        "codes": plan["codes"],
        "points_by_code": plan["points_by_code"],
        "points_per_code": plan.get("points_per_code"),
        "total_points": total_points,
        "hits_by_code": hits,
        "hit_codes": [code for code, count in hits.items() if count],
        "hit_units": hit_units,
        "hit_day": hit_units > 0,
        "profit_day": pnl > 0,
        "capital_vnd": capital,
        "payout_vnd": payout,
        "pnl_vnd": pnl,
        "roi_pct": 100 * pnl / capital,
        "draw_hash": digest(draw),
    }


def update_period(period: dict, settlement: dict) -> dict:
    value = dict(period)
    value["sessions"] += 1
    value["hit_days"] += int(settlement["hit_day"])
    value["profit_days"] += int(settlement["profit_day"])
    value["capital_vnd"] += settlement["capital_vnd"]
    value["payout_vnd"] += settlement["payout_vnd"]
    value["net_profit_vnd"] += settlement["pnl_vnd"]
    value["equity_current_vnd"] += settlement["pnl_vnd"]
    value["equity_peak_vnd"] = max(
        value["equity_peak_vnd"], value["equity_current_vnd"]
    )
    current_dd = value["equity_current_vnd"] - value["equity_peak_vnd"]
    value["max_drawdown_vnd"] = min(value["max_drawdown_vnd"], current_dd)
    tracks_gross = (
        value.get("gross_profit_vnd") is not None
        and value.get("gross_loss_vnd") is not None
    )
    if tracks_gross:
        value["gross_profit_vnd"] += max(0, settlement["pnl_vnd"])
        value["gross_loss_vnd"] += max(0, -settlement["pnl_vnd"])
    value["hit_day_rate_pct"] = 100 * value["hit_days"] / value["sessions"]
    value["profit_day_rate_pct"] = (
        100 * value["profit_days"] / value["sessions"]
    )
    value["roi_pct"] = 100 * value["net_profit_vnd"] / value["capital_vnd"]
    if tracks_gross:
        loss = value["gross_loss_vnd"]
        value["profit_factor"] = value["gross_profit_vnd"] / loss if loss else None
    else:
        value["profit_factor"] = None
    return value


def empty_period() -> dict:
    return {
        "sessions": 0,
        "hit_days": 0,
        "hit_day_rate_pct": 0.0,
        "profit_days": 0,
        "profit_day_rate_pct": 0.0,
        "capital_vnd": 0,
        "payout_vnd": 0,
        "net_profit_vnd": 0,
        "roi_pct": 0.0,
        "profit_factor": None,
        "max_drawdown_vnd": 0,
        "equity_current_vnd": 0,
        "equity_peak_vnd": 0,
        "gross_profit_vnd": 0,
        "gross_loss_vnd": 0,
    }


def advance_state(state: dict, settlement: dict) -> dict:
    value = json.loads(json.dumps(state))
    value["full"] = update_period(value["full"], settlement)
    settled_day = date.fromisoformat(settlement["date"])
    month_id = settled_day.strftime("%Y-%m")
    if value.get("current_month_id") != month_id:
        value["current_month_id"] = month_id
        value["current_month"] = empty_period()
    value["current_month"] = update_period(value["current_month"], settlement)
    value["settled_through"] = settled_day.isoformat()
    value["last_settlement_hash"] = digest(settlement)
    return value


def run_engine(source_end: date, workbook: Path) -> dict:
    command = [
        sys.executable,
        str(ENGINE / "runtime_plan.py"),
        "--source-end",
        source_end.isoformat(),
    ]
    environment = dict(os.environ)
    environment["MB_V32_SOURCE_XLSX"] = str(workbook)
    environment["MB_V32_CODE_CACHE"] = str(workbook.parent / "v9_code_predictions.pkl")
    environment["MB_V32_SEED_CACHE"] = str(
        workbook.parent / "v9_code_seed_predictions.pkl"
    )
    completed = subprocess.run(
        command,
        cwd=ROOT,
        env=environment,
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        timeout=25 * 60,
        check=False,
    )
    marker = "V32_PLAN_JSON="
    records = [line[len(marker):] for line in completed.stdout.splitlines()
               if line.startswith(marker)]
    if completed.returncode or len(records) != 1:
        tail = "\n".join(completed.stdout.splitlines()[-20:])
        raise PipelineError(f"Engine V32 thất bại ({completed.returncode}):\n{tail}")
    return json.loads(records[0])


def engine_manifest_hash() -> str:
    if not MANIFEST.exists():
        raise PipelineError("Thiếu engine_v32/manifest.json")
    manifest = read_json(MANIFEST)
    files = sorted(
        path for path in ENGINE.iterdir()
        if path.is_file() and path.suffix in {".py", ".pkl"}
    )
    lines = [
        f"{sha256(path.read_bytes()).hexdigest()}  {path.name}\n"
        for path in files
    ]
    actual_tree = sha256("".join(lines).encode("utf-8")).hexdigest()
    if len(files) != manifest.get("file_count"):
        raise PipelineError("Số file engine khác manifest")
    if actual_tree != manifest.get("tree_sha256"):
        raise PipelineError("Tree hash engine khác manifest")
    return digest(manifest)


def make_plan(engine: dict, target: date, source_end: date) -> tuple[dict, dict]:
    raw = engine["plan"]
    if target != source_end + timedelta(days=1):
        raise PipelineError("Ngày mục tiêu phải ngay sau ngày khóa")
    if raw["date"] != target.isoformat() or engine["source_end"] != source_end.isoformat():
        raise PipelineError("Engine trả sai ngày mục tiêu hoặc ngày khóa")
    codes = [code2(value) for value in raw["codes"]]
    if not isinstance(raw.get("points"), dict):
        raise PipelineError("Engine thiếu mapping điểm")
    raw_points = {code2(key): int(value) for key, value in raw["points"].items()}
    width = len(codes)
    expected = 30 if width <= 6 else 25 if width <= 8 else 20
    if not 1 <= width <= 12:
        raise PipelineError("Engine vi phạm độ rộng 1-12")
    if len(codes) != len(set(codes)) or set(raw_points) != set(codes):
        raise PipelineError("Engine vi phạm mã duy nhất hoặc mapping điểm")
    if int(raw["points_per_code"]) != expected:
        raise PipelineError("Engine vi phạm bậc điểm A6_B8")
    if any(value != expected for value in raw_points.values()):
        raise PipelineError("Engine trả điểm âm/hỗn hợp hoặc sai A6_B8")
    points = raw_points
    if int(raw.get("N", -1)) != width:
        raise PipelineError("Engine trả N sai độ rộng")
    if int(raw.get("capital_vnd", -1)) != width * expected * COST_PER_POINT:
        raise PipelineError("Engine trả vốn sai công thức")
    if raw.get("outcome_known_at_selection") is not False:
        raise PipelineError("Engine không xác nhận khóa outcome mục tiêu")
    overlay_raw = raw["overlay"]
    margin = float(overlay_raw["normalized_margin"])
    threshold = float(overlay_raw["threshold"])
    if not math.isfinite(margin) or not math.isfinite(threshold) or threshold <= 0:
        raise PipelineError("Biên/ngưỡng overlay không hữu hạn hoặc ngưỡng <= 0")
    eligible = bool(overlay_raw["eligible_width"])
    active = bool(overlay_raw["active"])
    if active != (eligible and margin >= threshold):
        raise PipelineError("Trạng thái overlay không khớp ngưỡng")
    base_codes = {code2(value) for value in raw.get("base_codes", [])}
    parent_codes = {code2(value) for value in raw.get("parent_codes", [])}
    if not base_codes or not base_codes.issubset(set(codes)):
        raise PipelineError("Engine không bảo toàn tập base")
    if len(set(codes) - base_codes) > 4:
        raise PipelineError("Engine vượt giới hạn 4 mã online")
    if len(parent_codes) != width or not base_codes.issubset(parent_codes):
        raise PipelineError("Tập parent/base của engine không hợp lệ")
    proposed_remove = code2(overlay_raw["proposed_remove"])
    proposed_add = code2(overlay_raw["proposed_add"])
    if proposed_remove not in parent_codes:
        raise PipelineError("Mã đề xuất bỏ không thuộc parent")
    if active and (proposed_remove in codes or proposed_add not in codes):
        raise PipelineError("Engine báo swap active nhưng danh sách cuối không khớp")
    if not active and set(codes) != parent_codes:
        raise PipelineError("Engine báo giữ parent nhưng danh sách cuối đã đổi")
    total_points = sum(points.values())
    plan = {
        "target_date": target.isoformat(),
        "data_lock_date": source_end.isoformat(),
        "status": "LOCKED_WAITING_RESULT",
        "codes": codes,
        "number_of_codes": width,
        "points_per_code": expected,
        "points_by_code": points,
        "total_points": total_points,
        "cost_per_point_vnd": COST_PER_POINT,
        "total_capital_vnd": total_points * COST_PER_POINT,
        "maximum_loss_vnd": total_points * COST_PER_POINT,
        "core_100_enabled": False,
        "other_50_enabled": False,
        "outcome_known_at_selection": False,
    }
    overlay = {
        "rule": overlay_raw["rule"],
        "eligible_width": eligible,
        "proposed_remove": proposed_remove,
        "proposed_add": proposed_add,
        "normalized_margin": margin,
        "activation_threshold": threshold,
        "active": active,
        "decision": "APPLY_SWAP" if active else "KEEP_PARENT_PLAN",
        "reason": (
            "Đủ ngưỡng khóa; áp dụng đúng một thay đổi online."
            if active else "Biên đề xuất thấp hơn ngưỡng khóa hoặc độ rộng không hợp lệ; giữ danh sách cha."
        ),
    }
    return plan, overlay


def parse_person_date(value: Any) -> date | None:
    return as_date(value)


def parse_person_codes(value: Any) -> list[str]:
    if isinstance(value, bool):
        return []
    if isinstance(value, int):
        return [f"{value:02d}"] if 0 <= value <= 99 else []
    if isinstance(value, float):
        if not math.isfinite(value) or value < 0:
            return []
        if value.is_integer():
            integer = int(value)
            return [f"{integer:02d}"] if 0 <= integer <= 99 else []
        text = str(value)
    else:
        text = str(value or "").strip()
    if not text:
        return []
    if not re.fullmatch(r"\d{1,2}(?:\s*[,;.\-]\s*\d{1,2})*", text):
        return []
    codes = [f"{int(token):02d}" for token in re.split(r"\s*[,;.\-]\s*", text)]
    return codes if len(codes) == len(set(codes)) else []


def parse_total_points(value: Any) -> int | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value if value > 0 else None
    if isinstance(value, float):
        return int(value) if math.isfinite(value) and value > 0 and value.is_integer() else None
    text = str(value or "").strip()
    return int(text) if re.fullmatch(r"[1-9]\d*", text) else None


def personal_operations(snapshot: dict, day: date, draw: list[str]) -> tuple[list[dict], list[dict]]:
    counts = Counter(draw)
    operations, blocked = [], []
    entries = snapshot.get("people", [])
    names = [entry.get("name") for entry in entries]
    if len(names) != len(set(names)) or set(names) != set(PEOPLE):
        raise PipelineError("Snapshot cá nhân phải có đúng 5 slot duy nhất")
    by_name = {entry.get("name"): entry for entry in entries}
    for name in PEOPLE:
        entry = by_name.get(name)
        if not entry:
            blocked.append({"name": name, "reason": "MISSING_SNAPSHOT"})
            continue
        matched = False
        for offset, row in enumerate(entry.get("rows", []), start=1):
            padded = list(row) + [None] * (8 - len(row))
            if parse_person_date(padded[0]) != day:
                continue
            if str(padded[1] or "").strip().lower() == "số dư đầu kỳ":
                continue
            matched = True
            codes = parse_person_codes(padded[2])
            total_points = parse_total_points(padded[3])
            if not codes or not total_points or total_points % len(codes):
                blocked.append({
                    "name": name,
                    "sheet_name": entry.get("sheet_name", name),
                    "row": offset,
                    "reason": "AMBIGUOUS_CODES_OR_POINT_SPLIT",
                })
                continue
            points_each = total_points // len(codes)
            hit_units = sum(counts.get(code, 0) for code in codes)
            capital = total_points * COST_PER_POINT
            payout = hit_units * points_each * PAYOUT_PER_HIT_POINT
            pnl = payout - capital
            if padded[4] not in (None, ""):
                try:
                    current_pnl = int(float(padded[4]))
                except (TypeError, ValueError):
                    current_pnl = None
                if current_pnl != pnl:
                    blocked.append({
                        "name": name,
                        "sheet_name": entry.get("sheet_name", name),
                        "row": offset,
                        "reason": "EXISTING_PNL_CONFLICT",
                        "expected_pnl_vnd": pnl,
                    })
                    continue
            capital_text = f"{capital:,}".replace(",", ".")
            payout_text = f"{payout:,}".replace(",", ".")
            pnl_text = f"{pnl:+,}".replace(",", ".")
            note = (
                f"Tự động V32 06:00: {', '.join(codes)}; {points_each} điểm/số; "
                f"{hit_units} nháy; vốn {capital_text}đ; trả {payout_text}đ; "
                f"P/L {pnl_text}đ."
            )
            core = {
                "name": name,
                "sheet_name": entry.get("sheet_name", name),
                "row": offset,
                "date": day.isoformat(),
                "codes": codes,
                "total_points": total_points,
                "points_per_code": points_each,
                "pnl_vnd": pnl,
                "note": note,
                "expected_a_to_d_hash": digest(padded[:4]),
            }
            operations.append({
                "kind": "UPDATE_PERSONAL_PNL_IF_BLANK",
                "operation_id": digest(core),
                **core,
            })
        if not matched:
            core = {
                "kind": "LOG_PERSONAL_NO_ORDER",
                "name": name,
                "sheet_name": entry.get("sheet_name"),
                "date": day.isoformat(),
                "status": "NO_ORDER_ROW_PNL_NOT_INFERRED",
            }
            operations.append({"operation_id": digest(core), **core})
    return operations, blocked


def personal_input_fingerprint(snapshot: dict, day: date) -> str:
    """Hash only user-owned A:D inputs, stable after automation writes E/H."""
    entries = snapshot.get("people", [])
    names = [entry.get("name") for entry in entries]
    if len(names) != len(set(names)) or set(names) != set(PEOPLE):
        raise PipelineError("Snapshot cá nhân phải có đúng 5 slot duy nhất")
    by_name = {entry.get("name"): entry for entry in entries}
    material = []
    for name in PEOPLE:
        exists = name in by_name
        entry = by_name.get(name, {})
        rows = []
        for offset, row in enumerate(entry.get("rows", []), start=1):
            padded = list(row) + [None] * (4 - len(row))
            if parse_person_date(padded[0]) == day:
                rows.append({"row": offset, "a_to_d": padded[:4]})
        material.append({
            "name": name,
            "exists": exists,
            "sheet_name": entry.get("sheet_name"),
            "rows": rows,
        })
    return digest(material)


def source_operation(kind: str, record: dict) -> dict:
    core = {"kind": kind, "record": record}
    return {
        "kind": kind,
        "operation_id": digest(core),
        "row_hash": digest(record),
        "record": record,
    }


def build_sheet_payload(
    target: date,
    plan: dict,
    settlement: dict,
    snapshot: dict,
    input_hash: str,
) -> dict:
    now = f"{target.isoformat()}T06:00:00+07:00"
    settlement_record = {
        "record_type": "SETTLEMENT",
        "date": settlement["date"],
        "status": settlement["status"],
        "method": METHOD_WITH_STAKE,
        "codes": settlement["codes"],
        "points_by_code": settlement["points_by_code"],
        "total_points": settlement["total_points"],
        "capital_vnd": settlement["capital_vnd"],
        "payout_vnd": settlement["payout_vnd"],
        "pnl_vnd": settlement["pnl_vnd"],
        "hit_units": settlement["hit_units"],
        "source_date": settlement["date"],
        "input_hash": input_hash,
        "created_at": now,
    }
    plan_record = {
        "record_type": "PLAN",
        "date": plan["target_date"],
        "status": plan["status"],
        "method": METHOD_WITH_STAKE,
        "codes": plan["codes"],
        "points_by_code": plan["points_by_code"],
        "total_points": plan["total_points"],
        "capital_vnd": plan["total_capital_vnd"],
        "payout_vnd": None,
        "pnl_vnd": None,
        "hit_units": None,
        "source_date": plan["data_lock_date"],
        "input_hash": input_hash,
        "created_at": now,
    }
    personal, blocked = personal_operations(
        snapshot, date.fromisoformat(settlement["date"]), snapshot["draw"]
    )
    operations = [
        source_operation("UPSERT_SOURCE_SETTLEMENT", settlement_record),
        *personal,
        source_operation("UPSERT_SOURCE_PLAN", plan_record),
    ]
    return {
        "schema_version": "MB_V32_SHEETS_TXN_V2",
        "private": True,
        "must_not_publish_to_pages": True,
        "target_date": target.isoformat(),
        "input_hash": input_hash,
        "operations": operations,
        "blocked_personal_rows": blocked,
        "payload_hash": digest(operations),
    }


def fmt_vnd(value: int, signed: bool = False) -> str:
    prefix = "+" if signed and value > 0 else ""
    return (prefix + f"{value:,}đ").replace(",", ".")


def fmt_million(value: int, signed: bool = False) -> str:
    number = value / 1_000_000
    prefix = "+" if signed and number > 0 else ""
    text = f"{number:,.3f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return prefix + text + "tr"


def pct(value: float, digits: int = 4) -> str:
    return f"{value:.{digits}f}%".replace(".", ",")


def render(current: dict) -> str:
    plan = current["plan"]
    overlay = current["overlay"]
    settlement = current["latest_settlement"]
    periods = current["backtest"]
    month = next(
        value for key, value in periods.items()
        if key.startswith("month_") or key.startswith("july_")
    )
    full = next(value for key, value in periods.items() if key.startswith("full_"))
    target = date.fromisoformat(plan["target_date"])
    locked = date.fromisoformat(plan["data_lock_date"])
    settled_day = date.fromisoformat(settlement["date"])
    codes_text = ", ".join(plan["codes"])
    point = plan["points_per_code"]
    meter = min(100.0, max(0.0, 100 * overlay["normalized_margin"] /
                           overlay["activation_threshold"]))
    active = overlay["active"]
    replacements = {
        "AUDIT_ID": current["audit_id"],
        "DATA_LOCK_DMY": locked.strftime("%d/%m/%Y"),
        "TARGET_WEEKDAY": WEEKDAYS[target.weekday()],
        "TARGET_DMY": target.strftime("%d/%m/%Y"),
        "TARGET_DM": target.strftime("%d/%m"),
        "PLAN_DESCRIPTION": (
            f"{len(plan['codes'])} số đồng mức {point} điểm; không Core100, "
            "không Other50 và không thay số sau khi khóa."
        ),
        "PLAN_COLUMNS": str(min(10, max(4, len(plan["codes"])))),
        "NUMBERS_HTML": "\n          ".join(
            f'<div class="number"><b>{escape(code)}</b><span>{plan["points_by_code"][code]} ĐIỂM</span></div>'
            for code in plan["codes"]
        ),
        "PLAN_WIDTH": str(len(plan["codes"])),
        "TOTAL_POINTS": str(plan["total_points"]),
        "CAPITAL_VND": fmt_vnd(plan["total_capital_vnd"]),
        "COPY_PLAN": escape(
            f"{target.strftime('%d/%m/%Y')}: {codes_text} — mỗi số {point} điểm — "
            f"tổng vốn {fmt_vnd(plan['total_capital_vnd'])}", quote=True
        ),
        "OVERLAY_STATUS": "Đã kích hoạt" if active else "Không kích hoạt",
        "AUTOMATION_STATUS": (
            "Sheets đã xác minh"
            if current.get("automation", {}).get("status")
            == "SHEETS_APPLIED_READBACK_VERIFIED_READY_TO_PUBLISH"
            else "Chờ credential"
        ),
        "SETTLEMENT_DMY": settled_day.strftime("%d/%m/%Y"),
        "SETTLEMENT_HIT_UNITS": str(settlement["hit_units"]),
        "SETTLEMENT_HIT_CODES": (
            "Mã trúng: " + ", ".join(settlement["hit_codes"])
            if settlement["hit_codes"] else "Không có mã trúng"
        ),
        "SETTLEMENT_CAPITAL": fmt_vnd(settlement["capital_vnd"]),
        "SETTLEMENT_TOTAL_POINTS": str(settlement["total_points"]),
        "SETTLEMENT_PAYOUT": fmt_vnd(settlement["payout_vnd"]),
        "SETTLEMENT_PNL_CLASS": "green" if settlement["pnl_vnd"] >= 0 else "red",
        "SETTLEMENT_PNL": fmt_vnd(settlement["pnl_vnd"], signed=True),
        "SETTLEMENT_ROI": pct(settlement["roi_pct"], 2),
        "JULY_PERIOD": f"01–{locked.strftime('%d/%m')}",
        "JULY_HIT_RATE": pct(month["hit_day_rate_pct"], 2),
        "JULY_HIT_DAYS": str(month["hit_days"]),
        "JULY_SESSIONS": str(month["sessions"]),
        "JULY_PROFIT_RATE": pct(month["profit_day_rate_pct"], 2),
        "JULY_PROFIT_DAYS": str(month["profit_days"]),
        "JULY_NET": fmt_million(month["net_profit_vnd"], signed=True),
        "JULY_NET_CLASS": "green" if month["net_profit_vnd"] >= 0 else "red",
        "JULY_CAPITAL": fmt_million(month["capital_vnd"]),
        "JULY_PF": (
            f"{month['profit_factor']:.4f}".replace(".", ",")
            if month.get("profit_factor") is not None else "∞"
        ),
        "JULY_ROI": pct(month["roi_pct"], 4),
        "JULY_ROI_CLASS": "green" if month["roi_pct"] >= 0 else "red",
        "JULY_MAXDD": fmt_million(month["max_drawdown_vnd"]),
        "OVERLAY_REMOVE": overlay["proposed_remove"],
        "OVERLAY_ADD": overlay["proposed_add"],
        "OVERLAY_DECISION_HTML": (
            f"ĐỔI {overlay['proposed_remove']}<br>SANG {overlay['proposed_add']}"
            if active else f"KHÔNG ĐỔI<br>GIỮ {overlay['proposed_remove']}"
        ),
        "OVERLAY_MARGIN": f"{overlay['normalized_margin']:.6f}".replace(".", ","),
        "OVERLAY_THRESHOLD": str(overlay["activation_threshold"]).replace(".", ","),
        "OVERLAY_METER_PCT": f"{meter:.2f}%",
        "OVERLAY_NOTE": overlay["reason"],
        "OVERLAY_TITLE": "đã kích hoạt" if active else "giữ nguyên bản cha",
        "TIER_30_CLASS": "active" if point == 30 else "",
        "TIER_25_CLASS": "active" if point == 25 else "",
        "TIER_20_CLASS": "active" if point == 20 else "",
        "TIER_30_TODAY": " · HÔM NAY" if point == 30 else "",
        "TIER_25_TODAY": " · HÔM NAY" if point == 25 else "",
        "TIER_20_TODAY": " · HÔM NAY" if point == 20 else "",
        "AUDIT_REPLAY_TEXT": (
            f"Lệnh sinh lại từ nguồn đã khóa đến {locked.strftime('%d/%m')}."
        ),
        "AUDIT_CAPITAL_TEXT": (
            f"{len(plan['codes'])} × {point} × 23.000đ = "
            f"{fmt_vnd(plan['total_capital_vnd'])}."
        ),
        "AUDIT_CAUSAL_TEXT": (
            f"Kết quả {target.strftime('%d/%m')} không tham gia chọn số."
        ),
        "AUDIT_OVERLAY_TEXT": (
            f"{str(round(overlay['normalized_margin'], 6)).replace('.', ',')} "
            f"{'≥' if active else '<'} "
            f"{str(overlay['activation_threshold']).replace('.', ',')} nên "
            f"{'overlay được áp dụng' if active else 'overlay đứng yên'}."
        ),
        "FULL_SESSIONS": str(full["sessions"]),
        "FULL_HIT_DAYS": str(full["hit_days"]),
        "FULL_HIT_RATE": pct(full["hit_day_rate_pct"], 4),
        "FULL_PROFIT_RATE": pct(full["profit_day_rate_pct"], 4),
        "FULL_NET": fmt_million(full["net_profit_vnd"], signed=True),
        "FULL_NET_CLASS": "green" if full["net_profit_vnd"] >= 0 else "red",
        "FULL_CAPITAL": fmt_million(full["capital_vnd"]),
        "FULL_ROI": pct(full["roi_pct"], 4),
        "FULL_ROI_CLASS": "green" if full["roi_pct"] >= 0 else "red",
        "FULL_MAXDD": fmt_million(full["max_drawdown_vnd"]),
    }
    html = TEMPLATE.read_text(encoding="utf-8")
    for key, value in replacements.items():
        html = html.replace("{{" + key + "}}", str(value))
    leftovers = sorted(set(re.findall(r"\{\{[A-Z0-9_]+\}\}", html)))
    if leftovers:
        raise PipelineError(f"Template còn biến chưa thay: {leftovers}")
    return html


def public_payload(
    target: date,
    plan: dict,
    overlay: dict,
    settlement: dict,
    state: dict,
    engine: dict,
    engine_hash: str,
    input_hash: str,
) -> dict:
    through = settlement["date"].replace("-", "_")
    month_key = (
        f"month_{settlement['date'][:7].replace('-', '_')}_"
        f"through_{settlement['date'][-2:]}"
    )
    return {
        "schema_version": "MB_DAILY_WEB_V32_A6B8",
        "audit_id": f"MB-{target.strftime('%Y%m%d')}-V32-A6B8",
        "generated_at": f"{target.isoformat()}T06:00:00+07:00",
        "timezone": "Asia/Bangkok",
        "method": {
            "short_name": "V32 · A6_B8",
            "official_name": METHOD,
            "production_module": "engine_v32/production_v32_empirical_overdue_swap_r868.py",
            "status": "PRODUCTION_OFFICIAL",
            "rollback_parent": "R868Y26 + Residual Gate315",
            "coverage": "30/30",
            "promotion_note": "Chỉ thay khi nhánh mới vượt kiểm định nhân quả và ngoài mẫu.",
        },
        "plan": plan,
        "overlay": overlay,
        "stake_policy": [
            {"condition": "N_LE_6", "points_per_code": 30},
            {"condition": "N_7_TO_8", "points_per_code": 25},
            {"condition": "N_GE_9", "points_per_code": 20},
        ],
        "latest_settlement": settlement,
        "backtest": {
            month_key: state["current_month"],
            f"full_through_{through}": state["full"],
        },
        "constraints": {
            "maximum_total_codes": 12,
            "maximum_online_codes": 4,
            "duplicate_codes": 0,
            "preserve_base": True,
            "same_day_outcome_leakage": False,
        },
        "audit": {
            "input_hash": input_hash,
            "engine_manifest_hash": engine_hash,
            "plan_replayed_from_locked_source": True,
            "capital_formula_verified": True,
            "overlay_threshold_verified": True,
            "target_outcome_used": False,
            "causality": engine["causality"],
            "disclaimer": "Backtest là số liệu lịch sử, không bảo đảm kết quả tương lai.",
        },
        "automation": {
            "pipeline_version": PIPELINE_VERSION,
            "schedule": "06:00 Asia/Bangkok",
            "status": "PREPARED_SHEETS_REQUIRED_BEFORE_PUBLISH",
        },
    }


def prepare(args: argparse.Namespace) -> None:
    target = date.fromisoformat(args.target_date)
    previous = target - timedelta(days=1)
    output = Path(args.output).resolve()
    source = Path(args.source_xlsx).resolve()
    snapshot = read_json(Path(args.pnl_snapshot))
    history = load_history(source)
    if previous not in history:
        raise PipelineError(f"Chưa có kỳ {previous} đủ 27/27")
    previous_plan = load_plan(previous)
    engine_hash = engine_manifest_hash()
    prefix_material = [
        [day.isoformat(), history[day]] for day in sorted(history) if day <= previous
    ]
    state_before = read_json(STATE_FILE)
    immutable = {
        "pipeline_version": PIPELINE_VERSION,
        "target_date": target.isoformat(),
        "previous_plan_hash": digest(previous_plan),
        "history_prefix_hash": digest(prefix_material),
        "state_before_hash": digest(state_before),
        "personal_inputs_hash": personal_input_fingerprint(snapshot, previous),
        "sheet_binding_hash": digest({
            "source_sheet_id": snapshot.get("source_sheet_id"),
            "pnl_sheet_id": snapshot.get("pnl_sheet_id"),
            "source_crosscheck": snapshot.get("source_crosscheck"),
        }),
        "template_hash": sha256(TEMPLATE.read_bytes()).hexdigest(),
        "pipeline_code_hash": sha256(Path(__file__).read_bytes()).hexdigest(),
        "engine_manifest_hash": engine_hash,
    }
    input_hash = digest(immutable)
    committed_path = TX_DIR / f"{target.isoformat()}.json"
    if committed_path.exists():
        committed = read_json(committed_path)
        if committed.get("status") != "COMMITTED":
            raise PipelineError("Giao dịch cùng ngày đã tồn tại với input khác")
        locked = committed.get("immutable", {})
        comparable = set(immutable) - {"state_before_hash"}
        if any(immutable[key] != locked.get(key) for key in comparable):
            raise PipelineError("Giao dịch đã commit nhưng nguồn/code hiện tại xung đột")
        input_hash = committed["input_hash"]
        write_json(output / "prepared.json", {
            "status": "NOOP_ALREADY_COMMITTED",
            "target_date": target.isoformat(),
            "input_hash": input_hash,
            "noop": True,
        })
        write_json(output / "private" / "sheets_payload.json", {
            "schema_version": "MB_V32_SHEETS_TXN_V2",
            "target_date": target.isoformat(),
            "input_hash": input_hash,
            "operations": [],
            "payload_hash": digest([]),
            "noop": True,
        })
        print("V32_PREPARE_NOOP_ALREADY_COMMITTED")
        return

    state = state_before
    expected_through = (previous - timedelta(days=1)).isoformat()
    if state.get("settled_through") != expected_through:
        raise PipelineError(
            f"State đang đến {state.get('settled_through')}, cần {expected_through}"
        )
    settlement = settle(previous_plan, history[previous])
    state = advance_state(state, settlement)

    engine_workbook = output / "private" / "engine-input.xlsx"
    write_truncated_engine_workbook(history, previous, engine_workbook)
    engine = run_engine(previous, engine_workbook)
    plan, overlay = make_plan(engine, target, previous)
    current = public_payload(
        target, plan, overlay, settlement, state, engine,
        engine_hash, input_hash,
    )
    html = render(current)
    if 'data-static-dashboard="1"' not in html or "MB_STATUS_SAFE_V1" not in html:
        raise PipelineError("Dashboard thiếu marker an toàn")
    snapshot["draw"] = history[previous]
    sheets = build_sheet_payload(
        target, plan, settlement, snapshot, input_hash
    )
    transaction = {
        "pipeline_version": PIPELINE_VERSION,
        "status": "PREPARED",
        "target_date": target.isoformat(),
        "previous_date": previous.isoformat(),
        "input_hash": input_hash,
        "immutable": immutable,
        "settlement_hash": digest(settlement),
        "plan_hash": digest(plan),
        "public_json_hash": digest(current),
        "public_html_hash": sha256(html.encode("utf-8")).hexdigest(),
        "state_after_hash": digest(state),
        "sheets_payload_hash": sheets["payload_hash"],
        "sheets_operation_ids": [
            operation["operation_id"] for operation in sheets["operations"]
        ],
        "prepared_at": datetime.now(VN).isoformat(timespec="seconds"),
    }
    repo = output / "repo"
    write_text(repo / "index.html", html)
    write_json(repo / "data" / "current.json", current)
    write_json(repo / "data" / "v32-state.json", state)
    write_json(repo / "data" / "v32-plans" / f"{target}.json", plan)
    write_json(repo / "data" / "v32-settlements" / f"{previous}.json", settlement)
    write_json(repo / "data" / "v32-transactions" / f"{target}.json", transaction)
    write_json(output / "private" / "sheets_payload.json", sheets)
    write_json(output / "prepared.json", {
        **transaction,
        "noop": False,
        "repo_files": sorted(
            str(path.relative_to(repo)) for path in repo.rglob("*") if path.is_file()
        ),
    })
    print(f"V32_PREPARED={target}")


def finalize(args: argparse.Namespace) -> None:
    output = Path(args.output).resolve()
    prepared = read_json(output / "prepared.json")
    receipt = read_json(Path(args.receipt))
    if prepared.get("noop"):
        if receipt.get("status") != "NOOP":
            raise PipelineError("Receipt NOOP không hợp lệ")
        if receipt.get("target_date") != prepared["target_date"]:
            raise PipelineError("Receipt NOOP sai ngày")
        if receipt.get("input_hash") != prepared["input_hash"]:
            raise PipelineError("Receipt NOOP sai input hash")
        print("V32_FINALIZE_NOOP")
        return
    if receipt.get("status") != "APPLIED_READBACK_VERIFIED":
        raise PipelineError("Google Sheets chưa được ghi và đọc lại đầy đủ")
    if receipt.get("payload_hash") != prepared["sheets_payload_hash"]:
        raise PipelineError("Receipt Google Sheets sai payload hash")
    if receipt.get("target_date") != prepared["target_date"]:
        raise PipelineError("Receipt Google Sheets sai ngày mục tiêu")
    if receipt.get("input_hash") != prepared["input_hash"]:
        raise PipelineError("Receipt Google Sheets sai input hash")
    if receipt.get("operation_ids") != prepared["sheets_operation_ids"]:
        raise PipelineError("Receipt Google Sheets thiếu/thừa/sai thứ tự operation")
    repo_stage = output / "repo"
    declared_files = set(prepared.get("repo_files", []))
    actual_files = {
        str(path.relative_to(repo_stage))
        for path in repo_stage.rglob("*") if path.is_file()
    }
    if not declared_files or actual_files != declared_files:
        raise PipelineError("Staging có file thiếu/thừa so với allowlist prepare")
    if any(path.is_symlink() for path in repo_stage.rglob("*")):
        raise PipelineError("Staging chứa symlink không được phép")
    tx_rel = Path("data") / "v32-transactions" / f"{prepared['target_date']}.json"
    tx_path = repo_stage / tx_rel
    transaction = read_json(tx_path)
    root_tx = ROOT / tx_rel
    if root_tx.exists():
        existing = read_json(root_tx)
        if existing.get("status") == "COMMITTED" and existing.get("input_hash") == prepared["input_hash"]:
            print("V32_FINALIZE_NOOP_ALREADY_ROOT_COMMITTED")
            return
        raise PipelineError("Root đã có transaction cùng ngày xung đột")
    if transaction.get("status") != "PREPARED" or transaction.get("input_hash") != prepared["input_hash"]:
        raise PipelineError("Transaction staging không khớp prepared")
    plan_path = repo_stage / "data" / "v32-plans" / f"{prepared['target_date']}.json"
    settlement_path = (
        repo_stage / "data" / "v32-settlements" /
        f"{transaction['previous_date']}.json"
    )
    current_path = repo_stage / "data" / "current.json"
    state_path = repo_stage / "data" / "v32-state.json"
    html_path = repo_stage / "index.html"
    private_payload = read_json(output / "private" / "sheets_payload.json")
    checks = {
        "plan_hash": digest(read_json(plan_path)),
        "settlement_hash": digest(read_json(settlement_path)),
        "public_json_hash": digest(read_json(current_path)),
        "public_html_hash": sha256(html_path.read_bytes()).hexdigest(),
        "state_after_hash": digest(read_json(state_path)),
        "sheets_payload_hash": digest(private_payload.get("operations", [])),
    }
    for key, actual in checks.items():
        if actual != prepared[key]:
            raise PipelineError(f"Staging bị thay đổi sau prepare: {key}")
    if render(read_json(current_path)).encode("utf-8") != html_path.read_bytes():
        raise PipelineError("HTML staging không tái tạo được từ current.json")
    current = read_json(current_path)
    current["automation"]["status"] = "SHEETS_APPLIED_READBACK_VERIFIED_READY_TO_PUBLISH"
    current["automation"]["sheets_receipt_hash"] = digest(receipt)
    write_json(current_path, current)
    write_text(repo_stage / "index.html", render(current))
    transaction.update({
        "status": "COMMITTED",
        "sheets_receipt_hash": digest(receipt),
        "sheets_operation_ids": receipt.get("operation_ids", []),
        "final_public_json_hash": digest(current),
        "final_public_html_hash": sha256(html_path.read_bytes()).hexdigest(),
        "committed_at": datetime.now(VN).isoformat(timespec="seconds"),
    })
    write_json(tx_path, transaction)
    copy_order = sorted(declared_files - {str(tx_rel)}) + [str(tx_rel)]
    for relative in copy_order:
        source = repo_stage / relative
        destination = ROOT / relative
        destination.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(source, destination)
    print(f"V32_FINALIZED={prepared['target_date']}")


def verify_seed(_: argparse.Namespace) -> None:
    state = read_json(STATE_FILE)
    current = read_json(DATA / "current.json")
    plan = load_plan(date.fromisoformat(current["plan"]["target_date"]))
    if state["settled_through"] != current["plan"]["data_lock_date"]:
        raise PipelineError("Seed state và current.json lệch ngày")
    if plan != current["plan"]:
        raise PipelineError("Seed plan và current.json lệch nội dung")
    engine_manifest_hash()
    print("V32_SEED_AND_ENGINE_MANIFEST_OK")


def main() -> None:
    parser = argparse.ArgumentParser()
    sub = parser.add_subparsers(dest="command", required=True)
    prepare_parser = sub.add_parser("prepare")
    prepare_parser.add_argument("--target-date", required=True)
    prepare_parser.add_argument("--source-xlsx", required=True)
    prepare_parser.add_argument("--pnl-snapshot", required=True)
    prepare_parser.add_argument("--output", required=True)
    prepare_parser.set_defaults(func=prepare)
    finalize_parser = sub.add_parser("finalize")
    finalize_parser.add_argument("--output", required=True)
    finalize_parser.add_argument("--receipt", required=True)
    finalize_parser.set_defaults(func=finalize)
    verify_parser = sub.add_parser("verify-seed")
    verify_parser.set_defaults(func=verify_seed)
    args = parser.parse_args()
    try:
        args.func(args)
    except PipelineError as exc:
        print(f"V32_PIPELINE_BLOCKED: {exc}", file=sys.stderr)
        raise SystemExit(2) from exc


if __name__ == "__main__":
    main()
